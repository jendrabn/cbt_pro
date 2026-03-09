from __future__ import annotations

from io import BytesIO
from pathlib import Path
from urllib.parse import urlparse

from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

from apps.core.services import get_branding_settings


def _filename_timestamp():
    return timezone.localtime().strftime("%Y%m%d_%H%M%S")


def _resolve_local_media_path(file_url: str) -> str:
    text = str(file_url or "").strip()
    if not text:
        return ""
    parsed = urlparse(text)
    path_value = parsed.path if parsed.scheme else text
    if path_value.startswith(("http://", "https://")):
        return ""

    media_url = getattr(settings, "MEDIA_URL", "/media/")
    media_root = Path(getattr(settings, "MEDIA_ROOT", "media"))

    relative_path = ""
    if path_value.startswith(media_url):
        relative_path = path_value[len(media_url):]
    elif path_value.startswith("/media/"):
        relative_path = path_value[len("/media/"):]
    elif path_value.startswith("/"):
        return ""
    else:
        relative_path = path_value

    candidate = (media_root / relative_path.lstrip("/")).resolve()
    try:
        candidate.relative_to(media_root.resolve())
    except ValueError:
        return ""
    if not candidate.exists():
        return ""
    return str(candidate)


def _branding_payload():
    branding = get_branding_settings()
    return {
        "institution_name": branding.get("institution_name") or getattr(settings, "CBT_SITE_NAME", "Sistem CBT"),
        "institution_type": branding.get("institution_type") or "",
        "institution_address": branding.get("institution_address") or "",
        "institution_logo_url": branding.get("institution_logo_url") or "",
    }


def export_results_to_xlsx(exam, rows):
    branding = _branding_payload()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Hasil Ujian"

    worksheet.append([branding["institution_name"]])
    worksheet.append([f"{branding['institution_type']} | {branding['institution_address']}".strip(" |")])
    worksheet.append([f"Laporan Hasil Ujian: {exam.title}"])
    worksheet.append([f"Dicetak: {timezone.localtime().strftime('%d %b %Y %H:%M:%S')}"])
    worksheet.append([])

    worksheet.append(
        [
            "Peringkat",
            "Nama Siswa",
            "Username",
            "Kelas",
            "Skor",
            "Persentase",
            "Status",
            "Waktu Pengerjaan",
            "Total Pelanggaran",
        ]
    )

    for row in rows:
        worksheet.append(
            [
                row["rank"],
                row["student_name"],
                row["student_username"],
                row["class_label"],
                row["total_score"],
                row["percentage"],
                row["status_label"],
                row["time_taken_human"],
                row["total_violations"],
            ]
        )

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=9)
    worksheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=9)

    for column_index in range(1, worksheet.max_column + 1):
        column_cells = worksheet.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=1,
            max_row=worksheet.max_row,
        )
        max_len = max(len(str(cell.value or "")) for cell in next(column_cells))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_len + 2, 42)

    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="hasil_ujian_{exam.id}_{_filename_timestamp()}.xlsx"'
    )
    return response


def export_results_to_csv(exam, rows):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="hasil_ujian_{exam.id}_{_filename_timestamp()}.csv"'
    )
    response.write("\ufeff")

    headers = [
        "Peringkat",
        "Nama Siswa",
        "Username",
        "Kelas",
        "Skor",
        "Persentase",
        "Status",
        "Waktu Pengerjaan",
        "Total Pelanggaran",
    ]
    response.write(",".join(headers) + "\n")

    def _csv_escape(value):
        return '"' + str(value).replace('"', '""') + '"'

    for row in rows:
        values = [
            row["rank"],
            row["student_name"],
            row["student_username"],
            row["class_label"],
            row["total_score"],
            row["percentage"],
            row["status_label"],
            row["time_taken_human"],
            row["total_violations"],
        ]
        escaped = [_csv_escape(value) for value in values]
        response.write(",".join(escaped) + "\n")
    return response


def export_results_to_pdf(exam, rows, summary):
    branding = _branding_payload()
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )

    styles = getSampleStyleSheet()
    elements = []

    logo_path = _resolve_local_media_path(branding["institution_logo_url"])
    if logo_path:
        try:
            logo_image = Image(logo_path, width=1.8 * cm, height=1.8 * cm)
            logo_image.hAlign = "LEFT"
            elements.append(logo_image)
            elements.append(Spacer(1, 0.12 * cm))
        except Exception:
            pass

    institution_line = branding["institution_name"]
    institution_meta = " | ".join([item for item in [branding["institution_type"], branding["institution_address"]] if item])
    if institution_line:
        elements.append(Paragraph(institution_line, styles["Heading3"]))
    if institution_meta:
        elements.append(Paragraph(institution_meta, styles["BodyText"]))
    elements.append(Spacer(1, 0.2 * cm))

    title = Paragraph(f"Laporan Hasil Ujian: {exam.title}", styles["Heading2"])
    meta = Paragraph(
        (
            f"Mata Pelajaran: {summary['subject_name']} | "
            f"Peserta: {summary['total_participants']} | "
            f"Lulus: {summary['passed_count']} | "
            f"Belum Lulus: {summary['failed_count']} | "
            f"Tingkat Kelulusan: {summary['pass_rate']}%"
        ),
        styles["BodyText"],
    )

    elements.extend([title, Spacer(1, 0.2 * cm), meta, Spacer(1, 0.35 * cm)])

    table_data = [
        [
            "Peringkat",
            "Nama Siswa",
            "Kelas",
            "Skor",
            "Persentase",
            "Status",
            "Waktu",
            "Pelanggaran",
        ]
    ]
    for row in rows:
        table_data.append(
            [
                row["rank"],
                row["student_name"],
                row["class_label"],
                f"{row['total_score']:.2f}",
                f"{row['percentage']:.2f}%",
                row["status_label"],
                row["time_taken_human"],
                row["total_violations"],
            ]
        )

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#ced4da")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
            ]
        )
    )
    elements.append(table)

    document.build(elements)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="hasil_ujian_{exam.id}_{_filename_timestamp()}.pdf"'
    )
    return response


def export_certificates_to_xlsx(certificates, teacher):
    branding = _branding_payload()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sertifikat"

    worksheet.append([branding["institution_name"]])
    worksheet.append([f"{branding['institution_type']} | {branding['institution_address']}".strip(" |")])
    worksheet.append(["Laporan Sertifikat Ujian"])
    worksheet.append([f"Guru: {teacher.get_full_name().strip() or teacher.username}"])
    worksheet.append([f"Dicetak: {timezone.localtime().strftime('%d %b %Y %H:%M:%S')}"])
    worksheet.append([])

    worksheet.append(
        [
            "Nomor Sertifikat",
            "Nama Siswa",
            "Username",
            "Ujian",
            "Skor Akhir",
            "Persentase",
            "Status",
            "Terbit",
            "Dicabut",
        ]
    )

    for cert in certificates:
        if cert.revoked_at or not cert.is_valid:
            status = "Dicabut"
        elif cert.pdf_generated_at:
            status = "Siap"
        else:
            status = "Diproses"

        worksheet.append(
            [
                cert.certificate_number,
                cert.student.get_full_name().strip() or cert.student.username if cert.student_id else "-",
                cert.student.username if cert.student_id else "-",
                cert.exam.title if cert.exam_id else "-",
                float(cert.final_score or 0),
                float(cert.final_percentage or 0),
                status,
                timezone.localtime(cert.issued_at).strftime("%d-%m-%Y %H:%M:%S") if cert.issued_at else "-",
                timezone.localtime(cert.revoked_at).strftime("%d-%m-%Y %H:%M:%S") if cert.revoked_at else "-",
            ]
        )

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=9)
    worksheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=9)
    worksheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=9)

    for column_index in range(1, worksheet.max_column + 1):
        column_cells = worksheet.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=1,
            max_row=worksheet.max_row,
        )
        max_len = max(len(str(cell.value or "")) for cell in next(column_cells))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_len + 2, 44)

    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="sertifikat_guru_{teacher.id}_{_filename_timestamp()}.xlsx"'
    )
    return response


def get_certificate_branding_payload():
    branding = _branding_payload()
    return {
        "institution_name": branding["institution_name"],
        "institution_type": branding["institution_type"],
        "institution_logo_url": branding["institution_logo_url"],
    }
