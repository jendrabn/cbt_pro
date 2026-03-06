from __future__ import annotations
from dataclasses import dataclass, field
from typing import Any

from django.contrib.auth import get_user_model
from openpyxl import load_workbook

User = get_user_model()

TEACHER_REQUIRED_HEADERS = {"first_name", "last_name", "email", "username"}
TEACHER_OPTIONAL_HEADERS = {"teacher_id", "subject_specialization", "phone_number", "is_active"}
TEACHER_ALL_HEADERS = TEACHER_REQUIRED_HEADERS | TEACHER_OPTIONAL_HEADERS

STUDENT_REQUIRED_HEADERS = {"first_name", "last_name", "email", "username", "student_id", "class_grade"}
STUDENT_OPTIONAL_HEADERS = {"phone_number", "is_active"}
STUDENT_ALL_HEADERS = STUDENT_REQUIRED_HEADERS | STUDENT_OPTIONAL_HEADERS


@dataclass
class ImportRowResult:
    row_number: int
    data: dict[str, Any] = field(default_factory=dict)
    status: str = "valid"
    error: str = ""
    username: str = ""
    email: str = ""


@dataclass
class ImportPreviewResult:
    total_rows: int = 0
    valid_rows: list[ImportRowResult] = field(default_factory=list)
    skip_rows: list[ImportRowResult] = field(default_factory=list)
    error_rows: list[ImportRowResult] = field(default_factory=list)

    @property
    def valid_count(self):
        return len(self.valid_rows)

    @property
    def skip_count(self):
        return len(self.skip_rows)

    @property
    def error_count(self):
        return len(self.error_rows)


def _normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def _validate_email(email):
    import re
    if not email:
        return "Email wajib diisi."
    email = str(email).strip().lower()
    pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    if not re.match(pattern, email):
        return "Format email tidak valid."
    return None


def _validate_username(username):
    if not username:
        return "Username wajib diisi."
    username = str(username).strip()
    if len(username) < 3:
        return "Username minimal 3 karakter."
    if len(username) > 150:
        return "Username maksimal 150 karakter."
    import re
    if not re.match(r"^[a-zA-Z0-9_]+$", username):
        return "Username hanya boleh mengandung huruf, angka, dan underscore."
    return None


def _parse_bool(value, default=True):
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    raw = str(value).strip().lower()
    return raw in {"1", "true", "ya", "yes", "y"}


def _clean_value(value):
    if value is None:
        return ""
    return str(value).strip()


def _collect_existing_accounts(candidates: list[dict[str, str]]) -> tuple[set[str], set[str]]:
    emails = [item["email"] for item in candidates if item.get("email")]
    usernames = [item["username"] for item in candidates if item.get("username")]
    existing_emails = set(User.objects.filter(email__in=emails).values_list("email", flat=True))
    existing_usernames = set(User.objects.filter(username__in=usernames).values_list("username", flat=True))
    return existing_emails, existing_usernames


class ExcelUserImporter:
    MAX_ROWS = 500
    MAX_FILE_SIZE_KB = 5 * 1024

    def __init__(self, role: str):
        if role not in ("teacher", "student"):
            raise ValueError("Role harus 'teacher' atau 'student'.")
        self.role = role
        self.required_headers = TEACHER_REQUIRED_HEADERS if role == "teacher" else STUDENT_REQUIRED_HEADERS
        self.all_headers = TEACHER_ALL_HEADERS if role == "teacher" else STUDENT_ALL_HEADERS

    def get_template_headers(self):
        if self.role == "teacher":
            return [
                "first_name",
                "last_name",
                "email",
                "username",
                "teacher_id",
                "subject_specialization",
                "phone_number",
                "is_active",
            ]
        return [
            "first_name",
            "last_name",
            "email",
            "username",
            "student_id",
            "class_grade",
            "phone_number",
            "is_active",
        ]

    def parse_file(self, uploaded_file) -> ImportPreviewResult:
        result = ImportPreviewResult()

        workbook = None
        try:
            workbook = load_workbook(uploaded_file, data_only=True)
        except Exception as exc:
            result.error_rows.append(
                ImportRowResult(
                    row_number=0,
                    status="error",
                    error=f"Gagal membaca file Excel: {exc}",
                )
            )
            return result

        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            result.error_rows.append(
                ImportRowResult(
                    row_number=0,
                    status="error",
                    error="File Excel kosong.",
                )
            )
            return result

        headers = [_normalize_header(h) for h in header_row]
        header_set = set(h for h in headers if h)

        missing_required = self.required_headers - header_set
        if missing_required:
            result.error_rows.append(
                ImportRowResult(
                    row_number=0,
                    status="error",
                    error=f"Header wajib tidak ditemukan: {', '.join(sorted(missing_required))}.",
                )
            )
            return result

        header_map = {}
        for idx, header in enumerate(headers):
            if header in self.all_headers:
                header_map[header] = idx

        candidates = []
        for row_values in sheet.iter_rows(min_row=2, max_row=self.MAX_ROWS + 1, values_only=True):
            if not any(value not in (None, "") for value in row_values):
                continue
            row_email = ""
            row_username = ""
            email_idx = header_map.get("email")
            username_idx = header_map.get("username")
            if email_idx is not None and email_idx < len(row_values):
                row_email = _clean_value(row_values[email_idx]).lower()
            if username_idx is not None and username_idx < len(row_values):
                row_username = _clean_value(row_values[username_idx])
            candidates.append({"email": row_email, "username": row_username})
        existing_emails, existing_usernames = _collect_existing_accounts(candidates)

        file_emails = {}
        file_usernames = {}

        for row_idx, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in row_values):
                continue

            result.total_rows += 1
            if result.total_rows > self.MAX_ROWS:
                result.error_rows.append(
                    ImportRowResult(
                        row_number=row_idx,
                        status="error",
                        error=f"Maksimal {self.MAX_ROWS} baris data.",
                    )
                )
                break

            row_result = self._parse_row(row_idx, row_values, header_map, existing_emails, existing_usernames, file_emails, file_usernames)

            if row_result.status == "valid":
                result.valid_rows.append(row_result)
            elif row_result.status == "skip":
                result.skip_rows.append(row_result)
            else:
                result.error_rows.append(row_result)

        return result

    def _parse_row(self, row_number, row_values, header_map, existing_emails, existing_usernames, file_emails, file_usernames):
        data = {}
        for header, idx in header_map.items():
            if idx < len(row_values):
                data[header] = row_values[idx]
        return self._validate_row_data(
            row_number=row_number,
            data=data,
            existing_emails=existing_emails,
            existing_usernames=existing_usernames,
            file_emails=file_emails,
            file_usernames=file_usernames,
        )

    def _validate_row_data(self, row_number, data, existing_emails, existing_usernames, file_emails, file_usernames):
        data = {k: v for k, v in data.items() if k in self.all_headers}

        row_result = ImportRowResult(row_number=row_number, data=data)

        first_name = _clean_value(data.get("first_name"))
        last_name = _clean_value(data.get("last_name"))
        email = _clean_value(data.get("email")).lower()
        username = _clean_value(data.get("username"))

        row_result.email = email
        row_result.username = username

        if not first_name:
            row_result.status = "error"
            row_result.error = "Nama depan wajib diisi."
            return row_result

        if not last_name:
            row_result.status = "error"
            row_result.error = "Nama belakang wajib diisi."
            return row_result

        email_error = _validate_email(email)
        if email_error:
            row_result.status = "error"
            row_result.error = email_error
            return row_result

        username_error = _validate_username(username)
        if username_error:
            row_result.status = "error"
            row_result.error = username_error
            return row_result

        if email in existing_emails:
            row_result.status = "skip"
            row_result.error = "Email sudah terdaftar di database."
            return row_result

        if username in existing_usernames:
            row_result.status = "skip"
            row_result.error = "Username sudah terdaftar di database."
            return row_result

        if email in file_emails:
            row_result.status = "skip"
            row_result.error = f"Email duplikat dengan baris {file_emails[email]}."
            return row_result

        if username in file_usernames:
            row_result.status = "skip"
            row_result.error = f"Username duplikat dengan baris {file_usernames[username]}."
            return row_result

        if self.role == "student":
            student_id = _clean_value(data.get("student_id"))
            class_grade = _clean_value(data.get("class_grade"))
            if not student_id:
                row_result.status = "error"
                row_result.error = "NIS wajib diisi untuk siswa."
                return row_result
            if not class_grade:
                row_result.status = "error"
                row_result.error = "Kelas wajib diisi untuk siswa."
                return row_result
            data["student_id"] = student_id
            data["class_grade"] = class_grade

        if self.role == "teacher":
            data["teacher_id"] = _clean_value(data.get("teacher_id"))
            data["subject_specialization"] = _clean_value(data.get("subject_specialization"))

        data["first_name"] = first_name
        data["last_name"] = last_name
        data["email"] = email
        data["username"] = username
        data["phone_number"] = _clean_value(data.get("phone_number"))
        data["is_active"] = _parse_bool(data.get("is_active"), default=True)

        file_emails[email] = row_number
        file_usernames[username] = row_number

        row_result.data = data
        row_result.status = "valid"
        return row_result

