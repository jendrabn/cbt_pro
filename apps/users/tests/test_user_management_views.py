from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from apps.accounts.models import User, UserImportLog, UserProfile


class UserManagementViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user(
            username="admin_um",
            email="admin.um@cbt.com",
            password="AdminPass123!",
            first_name="Admin",
            last_name="Utama",
            role="admin",
            is_active=True,
            is_staff=True,
        )
        cls.teacher = User.objects.create_user(
            username="teacher_um",
            email="teacher.um@cbt.com",
            password="TeacherPass123!",
            first_name="Guru",
            last_name="Satu",
            role="teacher",
            is_active=True,
        )

    def test_admin_can_access_user_list(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("user_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Manajemen Pengguna")

    def test_non_admin_forbidden_access_user_list(self):
        self.client.force_login(self.teacher)
        response = self.client.get(reverse("user_list"))
        self.assertEqual(response.status_code, 403)

    def test_admin_can_create_student_user(self):
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("user_create"),
            data={
                "first_name": "Andi",
                "last_name": "Pratama",
                "email": "andi.pratama@cbt.com",
                "username": "andi.pratama",
                "role": "student",
                "is_active": "on",
                "phone_number": "08123456789",
                "student_id": "NIS-3001",
                "class_grade": "XII IPA 3",
                "password": "StudentPass123!",
            },
        )
        self.assertEqual(response.status_code, 302)

        created_user = User.objects.get(username="andi.pratama")
        self.assertEqual(created_user.role, "student")
        self.assertTrue(created_user.check_password("StudentPass123!"))

        profile = UserProfile.objects.get(user=created_user)
        self.assertEqual(profile.student_id, "NIS-3001")
        self.assertEqual(profile.class_grade, "XII IPA 3")

    def test_admin_can_edit_user_without_changing_password(self):
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("user_edit", kwargs={"pk": self.teacher.pk}),
            data={
                "first_name": "Guru",
                "last_name": "Satu",
                "email": "teacher.um@cbt.com",
                "username": "teacher_um",
                "role": "teacher",
                "is_active": "on",
                "teacher_id": "NIP-1001",
                "subject_specialization": "Matematika",
                "password": "",
            },
        )
        self.assertEqual(response.status_code, 302)

        self.teacher.refresh_from_db()
        self.assertTrue(self.teacher.check_password("TeacherPass123!"))

    def test_admin_can_edit_user_and_change_password(self):
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("user_edit", kwargs={"pk": self.teacher.pk}),
            data={
                "first_name": "Guru",
                "last_name": "Satu",
                "email": "teacher.um@cbt.com",
                "username": "teacher_um",
                "role": "teacher",
                "is_active": "on",
                "teacher_id": "NIP-1001",
                "subject_specialization": "Matematika",
                "password": "ResetPass123!",
            },
        )
        self.assertEqual(response.status_code, 302)

        self.teacher.refresh_from_db()
        self.assertTrue(self.teacher.check_password("ResetPass123!"))

    def test_bulk_deactivate_users(self):
        user_a = User.objects.create_user(
            username="student_bulk_a",
            email="student.bulk.a@cbt.com",
            password="StudentBulk123!",
            first_name="Bulk",
            last_name="A",
            role="student",
            is_active=True,
        )
        user_b = User.objects.create_user(
            username="student_bulk_b",
            email="student.bulk.b@cbt.com",
            password="StudentBulk123!",
            first_name="Bulk",
            last_name="B",
            role="student",
            is_active=True,
        )

        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("user_list"),
            data={
                "action": "deactivate",
                "selected_ids": [str(user_a.id), str(user_b.id)],
            },
        )
        self.assertEqual(response.status_code, 302)

        user_a.refresh_from_db()
        user_b.refresh_from_db()
        self.assertFalse(user_a.is_active)
        self.assertFalse(user_b.is_active)

    def test_admin_can_reset_student_session(self):
        student = User.objects.create_user(
            username="student_reset_admin",
            email="student.reset.admin@cbt.com",
            password="StudentReset123!",
            first_name="Sesi",
            last_name="Admin",
            role="student",
            is_active=True,
        )
        student_client = Client()
        login_response = student_client.post(
            reverse("login"),
            {"username": student.username, "password": "StudentReset123!", "remember_me": "on"},
        )
        self.assertRedirects(login_response, reverse("student_dashboard"))

        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("user_reset_session", kwargs={"pk": student.pk}),
            data={"next": reverse("user_detail", kwargs={"pk": student.pk})},
        )
        self.assertEqual(response.status_code, 302)

        follow_up = student_client.get(reverse("student_dashboard"))
        self.assertEqual(follow_up.status_code, 302)
        self.assertTrue(follow_up.url.startswith(reverse("login")))

    def test_user_import_template_uses_machine_headers(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("user_import_template", kwargs={"role": "student"}))
        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        headers = [worksheet.cell(row=1, column=col_idx).value for col_idx in range(1, 9)]
        self.assertEqual(
            headers,
            [
                "first_name",
                "last_name",
                "email",
                "username",
                "student_id",
                "class_grade",
                "phone_number",
                "is_active",
            ],
        )

    def test_admin_can_import_users_from_excel_and_log_result(self):
        self.client.force_login(self.admin)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(
            [
                "first_name",
                "last_name",
                "email",
                "username",
                "student_id",
                "class_grade",
                "phone_number",
                "is_active",
            ]
        )
        worksheet.append(
            [
                "Rina",
                "Saputri",
                "rina.saputri@cbt.com",
                "rina.saputri",
                "NIS-9001",
                "XI IPA 1",
                "081111111111",
                "TRUE",
            ]
        )
        payload = BytesIO()
        workbook.save(payload)
        upload = SimpleUploadedFile(
            "import_users.xlsx",
            payload.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("user_import"),
            data={
                "import_file": upload,
                "role": "student",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(User.objects.filter(username="rina.saputri").exists())
        self.assertContains(response, "Password default untuk user hasil import batch ini adalah")
        self.assertContains(response, "Password default batch impor ini:")

        imported_user = User.objects.get(username="rina.saputri")
        self.assertTrue(imported_user.check_password(timezone.localtime().strftime("%Y%m%d")))

        import_log = UserImportLog.objects.latest("created_at")
        self.assertEqual(import_log.total_created, 1)
        self.assertEqual(import_log.total_failed, 0)
        self.assertEqual(import_log.status, "completed")
