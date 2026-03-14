from __future__ import annotations

from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.accounts.models import UserImportLog
from apps.subjects.models import Subject


class ImportTemplateExporter:
    @staticmethod
    def _append_guidance_sheet(workbook: Workbook, role: str):
        guide = workbook.create_sheet("Panduan")
        guide.append(["Kolom", "Wajib", "Aturan Nilai"])
        guide.append(["first_name", "YA", "Nama depan pengguna"])
        guide.append(["last_name", "YA", "Nama belakang pengguna"])
        guide.append(["email", "YA", "Gunakan format email yang valid dan unik"])
        guide.append(["username", "YA", "Huruf/angka/underscore, minimal 3 karakter, harus unik"])
        guide.append(["is_active", "TIDAK", "Isi TRUE atau FALSE"])
        guide.append(["phone_number", "TIDAK", "Nomor telepon (opsional)"])

        if role == "teacher":
            guide.append(["teacher_id", "TIDAK", "NIP guru (opsional)"])
            subject_names = list(Subject.objects.order_by("name").values_list("name", flat=True))
            subject_text = ", ".join(subject_names) if subject_names else "-"
            guide.append(["subject_specialization", "TIDAK", f"Gunakan nama subject yang tersedia: {subject_text}"])
        else:
            guide.append(["student_id", "YA", "NIS siswa"])
            guide.append(["class_grade", "YA", "Contoh: X IPA 1 / XI IPS 2"])

        for col_idx in range(1, 4):
            guide.column_dimensions[guide.cell(row=1, column=col_idx).column_letter].width = 44 if col_idx == 3 else 22

    @staticmethod
    def _build_template(headers: list[str], example_data: list[str], role: str) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Template Import"

        worksheet.append(headers)
        worksheet.append(example_data)

        for col_idx in range(1, len(headers) + 1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = 20

        ImportTemplateExporter._append_guidance_sheet(workbook, role=role)

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def create_teacher_template() -> bytes:
        headers = [
            "first_name",
            "last_name",
            "email",
            "username",
            "teacher_id",
            "subject_specialization",
            "phone_number",
            "is_active",
        ]
        example_data = [
            "Budi",
            "Santoso",
            "budi.santoso@sekolah.id",
            "budi.santoso",
            "NIP-1001",
            "Matematika",
            "081234567890",
            "TRUE",
        ]
        return ImportTemplateExporter._build_template(headers, example_data, role="teacher")

    @staticmethod
    def create_student_template() -> bytes:
        headers = [
            "first_name",
            "last_name",
            "email",
            "username",
            "student_id",
            "class_grade",
            "phone_number",
            "is_active",
        ]
        example_data = [
            "Ahmad",
            "Wijaya",
            "ahmad.wijaya@siswa.sekolah.id",
            "ahmad.wijaya",
            "NIS-2024001",
            "XII IPA 1",
            "081234567891",
            "TRUE",
        ]
        return ImportTemplateExporter._build_template(headers, example_data, role="student")


class ImportReportExporter:
    @staticmethod
    def create_report(import_log: UserImportLog, rows_data: list[dict]) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Laporan Import"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        headers = [
            "No",
            "Nama Depan",
            "Nama Belakang",
            "Email",
            "Username",
            "Status",
            "Keterangan",
        ]
        if import_log.imported_by.role == "teacher" or any("teacher_id" in row for row in rows_data):
            headers.insert(5, "NIP")
            headers.insert(6, "Mapel")
        else:
            headers.insert(5, "NIS")
            headers.insert(6, "Kelas")

        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col_idx in range(1, len(headers) + 1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = 18

        for row_idx, row_data in enumerate(rows_data, start=2):
            status = row_data.get("status", "unknown")
            fill = success_fill
            if status == "skip":
                fill = skip_fill
            elif status == "error":
                fill = error_fill

            col_idx = 1
            worksheet.cell(row=row_idx, column=col_idx, value=row_idx - 1)
            col_idx += 1
            worksheet.cell(row=row_idx, column=col_idx, value=row_data.get("first_name", ""))
            col_idx += 1
            worksheet.cell(row=row_idx, column=col_idx, value=row_data.get("last_name", ""))
            col_idx += 1
            worksheet.cell(row=row_idx, column=col_idx, value=row_data.get("email", ""))
            col_idx += 1
            worksheet.cell(row=row_idx, column=col_idx, value=row_data.get("username", ""))
            col_idx += 1

            if "teacher_id" in row_data or "student_id" not in row_data:
                worksheet.cell(row=row_idx, column=col_idx, value=row_data.get("teacher_id", ""))
                col_idx += 1
                worksheet.cell(row=row_idx, column=col_idx, value=row_data.get("subject_specialization", ""))
            else:
                worksheet.cell(row=row_idx, column=col_idx, value=row_data.get("student_id", ""))
                col_idx += 1
                worksheet.cell(row=row_idx, column=col_idx, value=row_data.get("class_grade", ""))
            col_idx += 1

            status_label = {"valid": "Berhasil", "skip": "Dilewati", "error": "Gagal"}.get(status, status)
            worksheet.cell(row=row_idx, column=col_idx, value=status_label)
            col_idx += 1
            worksheet.cell(row=row_idx, column=col_idx, value=row_data.get("error", ""))

            for c in range(1, len(headers) + 1):
                worksheet.cell(row=row_idx, column=c).fill = fill

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
