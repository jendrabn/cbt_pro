from __future__ import annotations

from datetime import datetime
from io import BytesIO
import json

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.subjects.models import Subject

from .importers import export_template_headers
from .models import QuestionOption


OPTION_LETTERS = [choice.value for choice in QuestionOption.OptionLetter]


def _safe_timestamp():
    return timezone.localtime().strftime("%Y%m%d_%H%M%S")


def _question_to_export_row(question):
    option_map = {option.option_letter: option for option in question.options.all()}
    ordering_items = [
        {"text": item.item_text, "order": int(item.correct_order)}
        for item in question.ordering_items.all().order_by("correct_order")
    ]
    matching_pairs = [
        {
            "prompt": pair.prompt_text,
            "answer": pair.answer_text,
            "order": int(pair.pair_order),
        }
        for pair in question.matching_pairs.all().order_by("pair_order")
    ]
    blank_answers = {}
    for blank in question.blank_answers.all().order_by("blank_number"):
        payload = blank.accepted_answers or []
        if blank.is_case_sensitive or blank.blank_points is not None:
            payload = {
                "accepted_answers": blank.accepted_answers or [],
                "is_case_sensitive": bool(blank.is_case_sensitive),
                "blank_points": float(blank.blank_points) if blank.blank_points is not None else None,
            }
        blank_answers[str(blank.blank_number)] = payload
    correct_option = ""
    correct_options = []
    option_columns = {}
    for letter in OPTION_LETTERS:
        option = option_map.get(letter)
        option_columns[f"option_{letter.lower()}"] = option.option_text if option else ""
        if option and option.is_correct:
            correct_options.append(letter)
            if not correct_option:
                correct_option = letter

    answer = getattr(question, "correct_answer", None)
    tags = [relation.tag.name for relation in question.questiontagrelation_set.all()]

    return {
        "id": str(question.id),
        "subject": question.subject.name if question.subject_id else "",
        "category": question.category.name if question.category_id else "",
        "question_type": question.question_type,
        "question_text": question.question_text,
        "difficulty_level": question.difficulty_level or "",
        "points": float(question.points),
        "explanation": question.explanation or "",
        "checkbox_scoring": question.checkbox_scoring or "",
        "ordering_items": json.dumps(ordering_items, ensure_ascii=False) if ordering_items else "",
        "matching_pairs": json.dumps(matching_pairs, ensure_ascii=False) if matching_pairs else "",
        "blank_answers": json.dumps(blank_answers, ensure_ascii=False) if blank_answers else "",
        "question_image_url": question.question_image_url or "",
        "audio_play_limit": question.audio_play_limit or "",
        "video_play_limit": question.video_play_limit or "",
        "allow_previous": question.allow_previous,
        "allow_next": question.allow_next,
        "force_sequential": question.force_sequential,
        "time_limit_seconds": question.time_limit_seconds or "",
        **option_columns,
        "correct_option": ",".join(correct_options) if question.question_type == "checkbox" else correct_option,
        "correct_options": ",".join(correct_options) if question.question_type == "checkbox" else "",
        "answer_text": answer.answer_text if answer else "",
        "keywords": ", ".join(answer.keywords or []) if answer else "",
        "is_case_sensitive": bool(answer.is_case_sensitive) if answer else False,
        "max_word_count": answer.max_word_count if answer and answer.max_word_count else "",
        "tags": ", ".join(tags),
        "is_active": question.is_active,
        "created_at": timezone.localtime(question.created_at).isoformat() if isinstance(question.created_at, datetime) else "",
        "updated_at": timezone.localtime(question.updated_at).isoformat() if isinstance(question.updated_at, datetime) else "",
    }


def export_questions_to_excel(queryset):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Bank Soal"

    headers = export_template_headers() + ["created_at", "updated_at"]
    worksheet.append(headers)

    for question in queryset:
        row = _question_to_export_row(question)
        worksheet.append([row.get(header, "") for header in headers])

    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="bank_soal_{_safe_timestamp()}.xlsx"'
    return response


def export_questions_to_csv(queryset):
    headers = export_template_headers() + ["created_at", "updated_at"]
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="bank_soal_{_safe_timestamp()}.csv"'
    response.write("\ufeff")

    def _csv_escape(value):
        return '"' + str(value).replace('"', '""') + '"'

    response.write(",".join(headers) + "\n")
    for question in queryset:
        row = _question_to_export_row(question)
        values = [row.get(header, "") for header in headers]
        escaped = [_csv_escape(value) for value in values]
        response.write(",".join(escaped) + "\n")
    return response


def export_import_template_excel():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Template Import"
    headers = export_template_headers()
    worksheet.append(headers)
    sample_row = {
        "subject": "Matematika",
        "category": "Aljabar",
        "question_type": "multiple_choice",
        "question_text": "Hasil dari 2 + 2 adalah ...",
        "difficulty_level": "easy",
        "points": 5,
        "explanation": "Operasi penjumlahan dasar.",
        "question_image_url": "",
        "audio_play_limit": "",
        "video_play_limit": "",
        "allow_previous": True,
        "allow_next": True,
        "force_sequential": False,
        "time_limit_seconds": "",
        "option_a": "3",
        "option_b": "4",
        "option_c": "5",
        "correct_option": "B",
        "correct_options": "",
        "checkbox_scoring": "",
        "ordering_items": "",
        "matching_pairs": "",
        "blank_answers": "",
        "answer_text": "",
        "keywords": "",
        "is_case_sensitive": False,
        "max_word_count": "",
        "tags": "aljabar, dasar",
        "is_active": True,
    }
    worksheet.append([sample_row.get(header, "") for header in headers])

    guide = workbook.create_sheet("Panduan")
    guide.append(["Kolom", "Wajib", "Aturan Nilai"])
    subject_names = list(Subject.objects.filter(is_active=True).order_by("name").values_list("name", flat=True))
    subject_text = ", ".join(subject_names) if subject_names else "-"
    guide_rows = [
        ("subject", "YA", f"Nama/kode subject aktif. Daftar subject: {subject_text}"),
        ("question_type", "YA", "multiple_choice, checkbox, ordering, matching, fill_in_blank, essay, short_answer"),
        ("question_text", "YA", "Teks soal"),
        ("difficulty_level", "TIDAK", "easy, medium, hard"),
        ("checkbox_scoring", "TIDAK", "all_or_nothing, partial, partial_no_penalty (khusus checkbox)"),
        ("ordering_items", "KONDISIONAL", 'JSON array untuk ordering. Contoh: [{"text":"Tahap 1","order":1},{"text":"Tahap 2","order":2}]'),
        ("matching_pairs", "KONDISIONAL", 'JSON array untuk matching. Contoh: [{"prompt":"Indonesia","answer":"Jakarta"}]'),
        ("blank_answers", "KONDISIONAL", 'JSON object untuk fill in blank. Contoh: {"1":["Soekarno","Ir. Soekarno"]}'),
        ("question_image_url", "TIDAK", "URL gambar soal jika ada"),
        ("audio_play_limit", "TIDAK", "Angka > 0, kosongkan jika tanpa batas audio"),
        ("video_play_limit", "TIDAK", "Angka > 0, kosongkan jika tanpa batas video"),
        ("allow_previous", "TIDAK", "TRUE atau FALSE"),
        ("allow_next", "TIDAK", "TRUE atau FALSE"),
        ("force_sequential", "TIDAK", "TRUE atau FALSE"),
        ("is_case_sensitive", "TIDAK", "TRUE atau FALSE (khusus short_answer)"),
        ("correct_option", "KONDISIONAL", "A sampai J untuk multiple_choice atau A,C untuk checkbox"),
        ("correct_options", "TIDAK", "Alternatif field checkbox: A,C"),
        ("answer_text", "KONDISIONAL", "Wajib untuk essay/short_answer"),
        ("is_active", "TIDAK", "TRUE atau FALSE"),
    ]
    for row in guide_rows:
        guide.append(list(row))
    for col_idx in range(1, 4):
        guide.column_dimensions[guide.cell(row=1, column=col_idx).column_letter].width = 48 if col_idx == 3 else 24

    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="template_import_bank_soal.xlsx"'
    return response


def export_import_report_excel(import_log) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Laporan Import"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    headers = ["No", "Baris Excel", "Status", "Keterangan"]
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    widths = [10, 14, 14, 60]
    for col_idx, width in enumerate(widths, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = width

    rows_data = []
    for error in import_log.error_details or []:
        rows_data.append(
            {
                "row": error.get("row", "-"),
                "status": "error",
                "message": error.get("error", ""),
            }
        )
    for skip in import_log.skip_details or []:
        rows_data.append(
            {
                "row": skip.get("row", "-"),
                "status": "skip",
                "message": skip.get("reason", ""),
            }
        )

    if not rows_data:
        rows_data.append(
            {
                "row": "-",
                "status": "success",
                "message": "Tidak ada baris yang gagal atau dilewati pada import ini.",
            }
        )

    status_labels = {
        "success": "Berhasil",
        "skip": "Dilewati",
        "error": "Gagal",
    }
    status_fills = {
        "success": success_fill,
        "skip": skip_fill,
        "error": error_fill,
    }

    for row_idx, row_data in enumerate(rows_data, start=2):
        fill = status_fills.get(row_data["status"], success_fill)
        worksheet.cell(row=row_idx, column=1, value=row_idx - 1)
        worksheet.cell(row=row_idx, column=2, value=row_data.get("row", "-"))
        worksheet.cell(row=row_idx, column=3, value=status_labels.get(row_data["status"], row_data["status"]))
        worksheet.cell(row=row_idx, column=4, value=row_data.get("message", ""))

        for col_idx in range(1, len(headers) + 1):
            worksheet.cell(row=row_idx, column=col_idx).fill = fill

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
