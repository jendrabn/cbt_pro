import shutil
import tempfile
import json
import shutil
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import ProgrammingError
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from PIL import Image

from apps.accounts.models import User
from apps.questions.models import Question, QuestionCategory, QuestionImportLog
from apps.questions.services import get_question_import_history
from apps.subjects.models import Subject


class QuestionBankViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.teacher = User.objects.create_user(
            username="teacher_qb",
            email="teacher.qb@cbt.com",
            password="TeacherPass123!",
            first_name="Guru",
            last_name="BankSoal",
            role="teacher",
            is_active=True,
        )
        cls.other_teacher = User.objects.create_user(
            username="teacher_qb_other",
            email="teacher.other@cbt.com",
            password="TeacherPass123!",
            first_name="Guru",
            last_name="Lain",
            role="teacher",
            is_active=True,
        )
        cls.student = User.objects.create_user(
            username="student_qb",
            email="student.qb@cbt.com",
            password="StudentPass123!",
            first_name="Siswa",
            last_name="BankSoal",
            role="student",
            is_active=True,
        )
        cls.subject = Subject.objects.create(name="Matematika", code="MAT")
        cls.category = QuestionCategory.objects.create(name="Aljabar", is_active=True)

    def _create_sample_question(self):
        question = Question.objects.create(
            created_by=self.teacher,
            subject=self.subject,
            category=self.category,
            question_type="multiple_choice",
            question_text="2 + 2 = ...",
            points=5,
            difficulty_level="easy",
            allow_previous=True,
            allow_next=True,
            force_sequential=False,
            is_active=True,
        )
        question.options.create(option_letter="A", option_text="3", is_correct=False, display_order=1)
        question.options.create(option_letter="B", option_text="4", is_correct=True, display_order=2)
        return question

    def _build_image_upload(self, name="diagram.png", size=(64, 64), color="#1d4ed8", image_format="PNG"):
        buffer = BytesIO()
        Image.new("RGB", size, color).save(buffer, format=image_format)
        return SimpleUploadedFile(
            name,
            buffer.getvalue(),
            content_type=f"image/{image_format.lower()}",
        )

    def test_teacher_can_access_question_list(self):
        self.client.force_login(self.teacher)
        response = self.client.get(reverse("question_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Manajemen Bank Soal")

    def test_question_create_form_shows_option_j_field(self):
        self.client.force_login(self.teacher)
        response = self.client.get(reverse("question_create"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Opsi J")

    def test_non_teacher_forbidden_question_list(self):
        self.client.force_login(self.student)
        response = self.client.get(reverse("question_list"))
        self.assertEqual(response.status_code, 403)

    def test_teacher_can_create_multiple_choice_question(self):
        self.client.force_login(self.teacher)
        response = self.client.post(
            reverse("question_create"),
            data={
                "subject": str(self.subject.id),
                "category": str(self.category.id),
                "question_type": "multiple_choice",
                "question_text": "Ibu kota Indonesia adalah ...",
                "points": "10",
                "difficulty_level": "easy",
                "explanation": "Jawaban yang benar adalah Jakarta.",
                "allow_previous": "on",
                "allow_next": "on",
                "is_active": "on",
                "option_a": "Bandung",
                "option_b": "Jakarta",
                "option_c": "",
                "option_d": "",
                "option_e": "",
                "correct_option": "B",
                "tags": "geografi, indonesia",
            },
        )
        self.assertEqual(response.status_code, 302)
        created = Question.objects.get(question_text__icontains="Ibu kota Indonesia")
        self.assertEqual(created.question_type, "multiple_choice")
        self.assertEqual(created.options.count(), 2)
        self.assertTrue(created.options.filter(option_letter="B", is_correct=True).exists())

    def test_teacher_can_create_multiple_choice_question_with_option_j(self):
        self.client.force_login(self.teacher)
        response = self.client.post(
            reverse("question_create"),
            data={
                "subject": str(self.subject.id),
                "category": str(self.category.id),
                "question_type": "multiple_choice",
                "question_text": "Pilih hasil 4 + 6.",
                "audio_play_limit": "2",
                "video_play_limit": "1",
                "points": "10",
                "difficulty_level": "easy",
                "allow_previous": "on",
                "allow_next": "on",
                "is_active": "on",
                "option_a": "8",
                "option_b": "9",
                "option_j": "10",
                "correct_option": "J",
            },
        )
        self.assertEqual(response.status_code, 302)
        created = Question.objects.get(question_text__icontains="4 + 6")
        self.assertEqual(created.options.count(), 3)
        self.assertEqual(created.audio_play_limit, 2)
        self.assertEqual(created.video_play_limit, 1)
        self.assertTrue(created.options.filter(option_letter="J", is_correct=True, option_text="10").exists())

    def test_teacher_can_create_checkbox_question(self):
        self.client.force_login(self.teacher)
        response = self.client.post(
            reverse("question_create"),
            data={
                "subject": str(self.subject.id),
                "category": str(self.category.id),
                "question_type": "checkbox",
                "question_text": "Pilih semua bilangan genap.",
                "points": "8",
                "difficulty_level": "easy",
                "allow_previous": "on",
                "allow_next": "on",
                "is_active": "on",
                "option_a": "1",
                "option_b": "2",
                "option_c": "3",
                "option_d": "4",
                "option_e": "",
                "correct_options": ["B", "D"],
                "checkbox_scoring": "partial_no_penalty",
            },
        )

        self.assertEqual(response.status_code, 302)
        created = Question.objects.get(question_text__icontains="bilangan genap")
        self.assertEqual(created.question_type, "checkbox")
        self.assertEqual(created.checkbox_scoring, "partial_no_penalty")
        self.assertEqual(created.options.filter(is_correct=True).count(), 2)
        self.assertTrue(created.options.filter(option_letter="B", is_correct=True).exists())
        self.assertTrue(created.options.filter(option_letter="D", is_correct=True).exists())

    def test_teacher_can_create_ordering_question(self):
        self.client.force_login(self.teacher)
        response = self.client.post(
            reverse("question_create"),
            data={
                "subject": str(self.subject.id),
                "category": str(self.category.id),
                "question_type": "ordering",
                "question_text": "Urutkan tahapan siklus air.",
                "points": "6",
                "difficulty_level": "medium",
                "allow_previous": "on",
                "allow_next": "on",
                "is_active": "on",
                "ordering_item_1": "Evaporasi",
                "ordering_item_2": "Kondensasi",
                "ordering_item_3": "Presipitasi",
            },
        )

        self.assertEqual(response.status_code, 302)
        created = Question.objects.get(question_text__icontains="siklus air")
        self.assertEqual(created.question_type, "ordering")
        self.assertEqual(created.ordering_items.count(), 3)
        self.assertEqual(
            list(created.ordering_items.order_by("correct_order").values_list("item_text", flat=True)),
            ["Evaporasi", "Kondensasi", "Presipitasi"],
        )
        self.assertFalse(created.options.exists())
        self.assertFalse(Question.objects.filter(id=created.id, correct_answer__isnull=False).exists())

    def test_teacher_can_create_matching_question(self):
        self.client.force_login(self.teacher)
        response = self.client.post(
            reverse("question_create"),
            data={
                "subject": str(self.subject.id),
                "category": str(self.category.id),
                "question_type": "matching",
                "question_text": "Pasangkan konsep dengan definisinya.",
                "points": "7",
                "difficulty_level": "medium",
                "allow_previous": "on",
                "allow_next": "on",
                "is_active": "on",
                "matching_prompt_1": "Massa",
                "matching_answer_1": "Jumlah materi dalam benda",
                "matching_prompt_2": "Gaya",
                "matching_answer_2": "Tarikan atau dorongan",
            },
        )

        self.assertEqual(response.status_code, 302)
        created = Question.objects.get(question_text__icontains="konsep dengan definisinya")
        self.assertEqual(created.question_type, "matching")
        self.assertEqual(created.matching_pairs.count(), 2)
        self.assertEqual(
            list(created.matching_pairs.order_by("pair_order").values_list("prompt_text", flat=True)),
            ["Massa", "Gaya"],
        )

    def test_teacher_can_create_fill_in_blank_question(self):
        self.client.force_login(self.teacher)
        response = self.client.post(
            reverse("question_create"),
            data={
                "subject": str(self.subject.id),
                "category": str(self.category.id),
                "question_type": "fill_in_blank",
                "question_text": "Air mendidih pada suhu {{1}} derajat {{2}}.",
                "points": "9",
                "difficulty_level": "easy",
                "allow_previous": "on",
                "allow_next": "on",
                "is_active": "on",
                "blank_accepted_answers_1": "100, 100.0",
                "blank_points_1": "5",
                "blank_accepted_answers_2": "Celsius, celcius",
                "blank_points_2": "4",
            },
        )

        self.assertEqual(response.status_code, 302)
        created = Question.objects.get(question_text__icontains="Air mendidih")
        self.assertEqual(created.question_type, "fill_in_blank")
        self.assertEqual(created.blank_answers.count(), 2)
        self.assertEqual(created.blank_answers.get(blank_number=1).accepted_answers, ["100", "100.0"])
        self.assertFalse(created.blank_answers.get(blank_number=2).is_case_sensitive)

    def test_teacher_can_create_short_answer_question_without_case_sensitive_or_max_words(self):
        self.client.force_login(self.teacher)
        response = self.client.post(
            reverse("question_create"),
            data={
                "subject": str(self.subject.id),
                "category": str(self.category.id),
                "question_type": "short_answer",
                "question_text": "Ibu kota Indonesia adalah ...",
                "points": "5",
                "difficulty_level": "easy",
                "allow_previous": "on",
                "allow_next": "on",
                "is_active": "on",
                "answer_text": "Jakarta",
                "keywords": "jakarta, ibu kota",
                "is_case_sensitive": "on",
                "max_word_count": "3",
            },
        )

        self.assertEqual(response.status_code, 302)
        created = Question.objects.get(created_by=self.teacher, question_type="short_answer", points=5)
        self.assertFalse(created.correct_answer.is_case_sensitive)
        self.assertIsNone(created.correct_answer.max_word_count)

    def test_teacher_can_create_multiple_choice_question_with_image_only_richtext(self):
        self.client.force_login(self.teacher)
        response = self.client.post(
            reverse("question_create"),
            data={
                "subject": str(self.subject.id),
                "category": str(self.category.id),
                "question_type": "multiple_choice",
                "question_text": '<p><img src="/media/questions/richtext/sample-question.png" width="280" height="140"></p>',
                "points": "10",
                "difficulty_level": "easy",
                "explanation": "<table><tr><th>Petunjuk</th></tr><tr><td>Perhatikan gambar pada soal.</td></tr></table>",
                "allow_previous": "on",
                "allow_next": "on",
                "is_active": "on",
                "option_a": '<p><img src="/media/questions/richtext/sample-option-a.png" width="120" height="120"></p>',
                "option_b": "<p>Gambar B</p>",
                "option_c": "",
                "option_d": "",
                "option_e": "",
                "correct_option": "B",
            },
        )
        self.assertEqual(response.status_code, 302)
        created = Question.objects.get(created_by=self.teacher, question_type="multiple_choice", points=10)
        self.assertIn("<img", created.question_text)
        self.assertTrue(created.options.filter(option_letter="A").exists())
        self.assertIn("<img", created.options.get(option_letter="A").option_text)

    def test_teacher_create_question_sanitizes_richtext_payload(self):
        self.client.force_login(self.teacher)
        response = self.client.post(
            reverse("question_create"),
            data={
                "subject": str(self.subject.id),
                "category": str(self.category.id),
                "question_type": "multiple_choice",
                "question_text": '<p>Konten aman</p><script>alert("x")</script><img src="https://example.com/q.png" onerror="alert(1)">',
                "points": "10",
                "difficulty_level": "easy",
                "explanation": '<div>Penjelasan</div><style>body{display:none}</style>',
                "allow_previous": "on",
                "allow_next": "on",
                "is_active": "on",
                "option_a": '<p>Opsi A</p><img src="https://example.com/a.png" onerror="alert(1)">',
                "option_b": "<p>Opsi B</p>",
                "correct_option": "B",
            },
        )

        self.assertEqual(response.status_code, 302)
        created = Question.objects.get(question_text__icontains="Konten aman")
        self.assertNotIn("<script", created.question_text.lower())
        self.assertNotIn("onerror", created.question_text.lower())
        self.assertIn("<img", created.question_text.lower())
        self.assertNotIn("<style", (created.explanation or "").lower())
        self.assertNotIn("onerror", created.options.get(option_letter="A").option_text.lower())

    def test_teacher_can_upload_richtext_image(self):
        temp_media_root = tempfile.mkdtemp(prefix="cbt_question_media_")
        self.addCleanup(lambda: shutil.rmtree(temp_media_root, ignore_errors=True))
        self.client.force_login(self.teacher)
        image_file = self._build_image_upload(size=(2600, 1300))

        with override_settings(MEDIA_ROOT=temp_media_root):
            response = self.client.post(
                reverse("question_richtext_upload"),
                data={"file": image_file},
            )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn("location", payload)
        self.assertIn("/media/questions/richtext/", payload["location"])

        relative_path = payload["location"].split("/media/", 1)[-1].lstrip("/")
        stored_path = Path(temp_media_root) / relative_path
        self.assertTrue(stored_path.exists())
        with Image.open(stored_path) as stored_image:
            self.assertLessEqual(stored_image.width, 1920)
            self.assertEqual(stored_image.width, 1920)

    def test_teacher_cannot_upload_richtext_file_with_fake_image_extension(self):
        temp_media_root = tempfile.mkdtemp(prefix="cbt_question_media_invalid_")
        self.addCleanup(lambda: shutil.rmtree(temp_media_root, ignore_errors=True))
        self.client.force_login(self.teacher)
        fake_image = SimpleUploadedFile(
            "fake-diagram.png",
            b"this is not a real png payload",
            content_type="image/png",
        )

        with override_settings(MEDIA_ROOT=temp_media_root):
            response = self.client.post(
                reverse("question_richtext_upload"),
                data={"file": fake_image},
            )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["error"], "Tipe file media tidak dapat dikenali.")

    def test_teacher_can_browse_uploaded_richtext_media(self):
        temp_media_root = tempfile.mkdtemp(prefix="cbt_question_browser_")
        self.addCleanup(lambda: shutil.rmtree(temp_media_root, ignore_errors=True))
        self.client.force_login(self.teacher)
        image_file = self._build_image_upload(name="diagram-browser.png", size=(800, 400))

        with override_settings(MEDIA_ROOT=temp_media_root):
            upload_response = self.client.post(
                reverse("question_richtext_upload"),
                data={"file": image_file},
            )
            self.assertEqual(upload_response.status_code, 200)

            richtext_dir = Path(temp_media_root) / "questions" / "richtext" / str(self.teacher.pk)
            richtext_dir.mkdir(parents=True, exist_ok=True)
            audio_path = richtext_dir / "sample-audio.mp3"
            audio_path.write_bytes(b"ID3test-audio")

            image_response = self.client.get(reverse("question_richtext_browser"), data={"kind": "image"})
            media_response = self.client.get(reverse("question_richtext_browser"), data={"kind": "media"})

        self.assertEqual(image_response.status_code, 200)
        self.assertEqual(media_response.status_code, 200)

        image_payload = image_response.json()
        media_payload = media_response.json()
        self.assertTrue(image_payload["items"])
        self.assertIn("image", {item["kind"] for item in image_payload["items"]})
        self.assertIn("audio", {item["kind"] for item in media_payload["items"]})

    def test_richtext_media_browser_is_scoped_per_teacher(self):
        temp_media_root = tempfile.mkdtemp(prefix="cbt_question_browser_scope_")
        self.addCleanup(lambda: shutil.rmtree(temp_media_root, ignore_errors=True))

        with override_settings(MEDIA_ROOT=temp_media_root):
            self.client.force_login(self.teacher)
            first_upload = self.client.post(
                reverse("question_richtext_upload"),
                data={"file": self._build_image_upload(name="teacher-one.png", size=(640, 320))},
            )
            self.assertEqual(first_upload.status_code, 200)
            self.assertIn(f"/{self.teacher.pk}/", first_upload.json()["location"])

            self.client.force_login(self.other_teacher)
            second_upload = self.client.post(
                reverse("question_richtext_upload"),
                data={"file": self._build_image_upload(name="teacher-two.png", size=(640, 320), color="#0f766e")},
            )
            self.assertEqual(second_upload.status_code, 200)
            self.assertIn(f"/{self.other_teacher.pk}/", second_upload.json()["location"])

            browser_response = self.client.get(reverse("question_richtext_browser"), data={"kind": "image"})

        self.assertEqual(browser_response.status_code, 200)
        items = browser_response.json()["items"]
        self.assertTrue(items)
        self.assertTrue(all(f"/{self.other_teacher.pk}/" in item["url"] for item in items))
        self.assertTrue(all(f"/{self.teacher.pk}/" not in item["url"] for item in items))

    def test_question_create_ignores_external_next_url(self):
        self.client.force_login(self.teacher)
        response = self.client.post(
            reverse("question_create"),
            data={
                "subject": str(self.subject.id),
                "category": str(self.category.id),
                "question_type": "multiple_choice",
                "question_text": "Planet terbesar adalah ...",
                "points": "5",
                "difficulty_level": "easy",
                "allow_previous": "on",
                "allow_next": "on",
                "is_active": "on",
                "option_a": "Bumi",
                "option_b": "Jupiter",
                "correct_option": "B",
                "next": "https://evil.example/phishing",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("question_list"))

    def test_question_edit_ignores_external_next_url(self):
        question = self._create_sample_question()
        self.client.force_login(self.teacher)
        response = self.client.post(
            reverse("question_edit", kwargs={"pk": question.pk}),
            data={
                "subject": str(self.subject.id),
                "category": str(self.category.id),
                "question_type": "multiple_choice",
                "question_text": "2 + 2 = ...",
                "points": "5",
                "difficulty_level": "easy",
                "allow_previous": "on",
                "allow_next": "on",
                "is_active": "on",
                "option_a": "3",
                "option_b": "4",
                "correct_option": "B",
                "next": "https://evil.example/phishing",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("question_list"))

    def test_question_preview_sanitizes_legacy_richtext(self):
        question = self._create_sample_question()
        Question.objects.filter(id=question.id).update(
            question_text='<p>Legacy aman</p><script>alert("x")</script><img src="https://example.com/legacy.png" onerror="alert(1)">',
            explanation='<div>Pembahasan</div><script>alert("x")</script>',
        )
        question.options.filter(option_letter="A").update(
            option_text='<p>Legacy opsi</p><img src="https://example.com/legacy-option.png" onerror="alert(1)"><script>alert("x")</script>'
        )

        self.client.force_login(self.teacher)
        response = self.client.get(reverse("question_preview", kwargs={"pk": question.pk}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Legacy aman")
        self.assertContains(response, "Legacy opsi")
        self.assertNotContains(response, 'alert("x")', html=False)
        self.assertNotContains(response, "onerror", html=False)

    def test_question_edit_matching_legacy_richtext_is_rendered_as_plain_text(self):
        question = Question.objects.create(
            created_by=self.teacher,
            subject=self.subject,
            category=self.category,
            question_type="matching",
            question_text="Pasangkan istilah dengan maknanya.",
            points=5,
            difficulty_level="easy",
            allow_previous=True,
            allow_next=True,
            force_sequential=False,
            is_active=True,
        )
        question.matching_pairs.create(prompt_text="<p>Massa</p>", answer_text="<div>Jumlah<br>materi</div>", pair_order=1)
        question.matching_pairs.create(prompt_text="<p>Gaya</p>", answer_text="<div>Tarikan</div>", pair_order=2)

        self.client.force_login(self.teacher)
        response = self.client.get(reverse("question_edit", kwargs={"pk": question.pk}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Massa")
        self.assertContains(response, "Jumlah")
        self.assertNotContains(response, "&lt;p&gt;Massa&lt;/p&gt;", html=False)
        self.assertNotContains(response, "&lt;div&gt;Jumlah&lt;br&gt;materi&lt;/div&gt;", html=False)

    def test_teacher_can_duplicate_question(self):
        question = self._create_sample_question()
        self.client.force_login(self.teacher)
        response = self.client.post(reverse("question_duplicate", kwargs={"pk": question.pk}))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Question.objects.filter(created_by=self.teacher, is_deleted=False).count(), 2)

    def test_teacher_can_open_question_preview(self):
        question = self._create_sample_question()
        self.client.force_login(self.teacher)
        response = self.client.get(reverse("question_preview", kwargs={"pk": question.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Pratinjau Soal")
        self.assertEqual(response.get("X-Frame-Options"), "SAMEORIGIN")

    def test_teacher_can_export_question_excel(self):
        self._create_sample_question()
        self.client.force_login(self.teacher)
        response = self.client.get(reverse("question_export"), data={"format": "xlsx"})
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        self.assertIn("attachment;", response["Content-Disposition"])

    def test_teacher_can_export_question_excel_with_option_j(self):
        question = Question.objects.create(
            created_by=self.teacher,
            subject=self.subject,
            category=self.category,
            question_type="multiple_choice",
            question_text="Pilih hasil 3 + 7.",
            points=5,
            difficulty_level="easy",
            allow_previous=True,
            allow_next=True,
            force_sequential=False,
            is_active=True,
        )
        question.options.create(option_letter="A", option_text="9", is_correct=False, display_order=1)
        question.options.create(option_letter="J", option_text="10", is_correct=True, display_order=2)
        question.audio_play_limit = 2
        question.video_play_limit = 1
        question.save(update_fields=["audio_play_limit", "video_play_limit"])

        self.client.force_login(self.teacher)
        response = self.client.get(reverse("question_export"), data={"format": "xlsx"})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        headers = list(rows[0])
        data_row = dict(zip(headers, rows[1]))

        self.assertIn("option_j", headers)
        self.assertEqual(data_row["option_j"], "10")
        self.assertEqual(data_row["correct_option"], "J")
        self.assertEqual(data_row["audio_play_limit"], 2)
        self.assertEqual(data_row["video_play_limit"], 1)

    def test_teacher_can_export_question_excel_with_ordering_items(self):
        question = Question.objects.create(
            created_by=self.teacher,
            subject=self.subject,
            category=self.category,
            question_type="ordering",
            question_text="Urutkan fase bulan.",
            points=5,
            difficulty_level="easy",
            allow_previous=True,
            allow_next=True,
            force_sequential=False,
            is_active=True,
        )
        question.ordering_items.create(item_text="Bulan Baru", correct_order=1)
        question.ordering_items.create(item_text="Kuartal Awal", correct_order=2)
        question.ordering_items.create(item_text="Purnama", correct_order=3)

        self.client.force_login(self.teacher)
        response = self.client.get(reverse("question_export"), data={"format": "xlsx"})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        headers = list(rows[0])
        data_rows = [dict(zip(headers, row)) for row in rows[1:]]
        export_row = next(row for row in data_rows if row["question_text"] == "Urutkan fase bulan.")

        self.assertIn("ordering_items", headers)
        self.assertIn('"Bulan Baru"', export_row["ordering_items"])
        self.assertIn('"order": 1', export_row["ordering_items"])

    def test_teacher_can_export_question_excel_with_matching_pairs(self):
        question = Question.objects.create(
            created_by=self.teacher,
            subject=self.subject,
            category=self.category,
            question_type="matching",
            question_text="Pasangkan besaran dengan satuannya.",
            points=5,
            difficulty_level="easy",
            allow_previous=True,
            allow_next=True,
            force_sequential=False,
            is_active=True,
        )
        question.matching_pairs.create(prompt_text="Panjang", answer_text="Meter", pair_order=1)
        question.matching_pairs.create(prompt_text="Massa", answer_text="Kilogram", pair_order=2)

        self.client.force_login(self.teacher)
        response = self.client.get(reverse("question_export"), data={"format": "xlsx"})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        headers = list(rows[0])
        data_rows = [dict(zip(headers, row)) for row in rows[1:]]
        export_row = next(row for row in data_rows if row["question_text"] == "Pasangkan besaran dengan satuannya.")

        self.assertIn("matching_pairs", headers)
        self.assertIn('"Panjang"', export_row["matching_pairs"])
        self.assertIn('"Meter"', export_row["matching_pairs"])

    def test_teacher_can_export_question_excel_with_fill_in_blank_answers(self):
        question = Question.objects.create(
            created_by=self.teacher,
            subject=self.subject,
            category=self.category,
            question_type="fill_in_blank",
            question_text="Planet terdekat dari Matahari adalah {{1}}.",
            points=5,
            difficulty_level="easy",
            allow_previous=True,
            allow_next=True,
            force_sequential=False,
            is_active=True,
        )
        question.blank_answers.create(blank_number=1, accepted_answers=["Merkurius", "Mercury"], blank_points=5)

        self.client.force_login(self.teacher)
        response = self.client.get(reverse("question_export"), data={"format": "xlsx"})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        headers = list(rows[0])
        data_rows = [dict(zip(headers, row)) for row in rows[1:]]
        export_row = next(row for row in data_rows if row["question_text"] == "Planet terdekat dari Matahari adalah {{1}}.")

        self.assertIn("blank_answers", headers)
        self.assertIn('"Merkurius"', export_row["blank_answers"])
        self.assertIn('"blank_points": 5', export_row["blank_answers"])

    def test_teacher_can_export_question_csv(self):
        self._create_sample_question()
        self.client.force_login(self.teacher)
        response = self.client.get(reverse("question_export"), data={"format": "csv"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        self.assertIn(".csv", response["Content-Disposition"])

    def test_teacher_can_bulk_delete_questions(self):
        question = self._create_sample_question()
        self.client.force_login(self.teacher)
        response = self.client.post(
            reverse("question_list"),
            data={
                "action": "delete",
                "selected_ids": [str(question.id)],
            },
        )
        self.assertEqual(response.status_code, 302)
        question.refresh_from_db()
        self.assertTrue(question.is_deleted)

    def test_teacher_can_import_question_excel(self):
        self.client.force_login(self.teacher)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(
            [
                "subject",
                "category",
                "question_type",
                "question_text",
                "difficulty_level",
                "points",
                "option_a",
                "option_b",
                "correct_option",
            ]
        )
        worksheet.append(
            [
                "Matematika",
                "Aljabar",
                "multiple_choice",
                "5 + 5 = ...",
                "easy",
                5,
                "9",
                "10",
                "B",
            ]
        )
        payload = BytesIO()
        workbook.save(payload)
        upload = SimpleUploadedFile(
            "import_soal.xlsx",
            payload.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(reverse("question_import"), data={"import_file": upload})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(Question.objects.filter(question_text__icontains="5 + 5").exists())
        self.assertContains(response, "Ringkasan Impor Terakhir")
        self.assertContains(response, "Template Bank Soal")
        self.assertContains(response, "Riwayat Impor")
        import_log = QuestionImportLog.objects.latest("created_at")
        self.assertEqual(import_log.total_created, 1)
        self.assertEqual(import_log.total_failed, 0)
        self.assertEqual(import_log.status, "completed")

        report_response = self.client.get(reverse("question_import_report", kwargs={"log_id": import_log.id}))
        self.assertEqual(report_response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            report_response["Content-Type"],
        )

    def test_teacher_can_import_question_excel_with_option_j(self):
        self.client.force_login(self.teacher)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(
            [
                "subject",
                "category",
                "question_type",
                "question_text",
                "difficulty_level",
                "points",
                "audio_play_limit",
                "video_play_limit",
                "option_a",
                "option_b",
                "option_j",
                "correct_option",
            ]
        )
        worksheet.append(
            [
                "Matematika",
                "Aljabar",
                "multiple_choice",
                "7 + 5 = ...",
                "easy",
                5,
                2,
                1,
                "10",
                "11",
                "12",
                "J",
            ]
        )
        payload = BytesIO()
        workbook.save(payload)
        upload = SimpleUploadedFile(
            "import_soal_option_j.xlsx",
            payload.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(reverse("question_import"), data={"import_file": upload})

        self.assertEqual(response.status_code, 200)
        created = Question.objects.get(question_text__icontains="7 + 5")
        self.assertEqual(created.audio_play_limit, 2)
        self.assertEqual(created.video_play_limit, 1)
        self.assertTrue(created.options.filter(option_letter="J", is_correct=True, option_text="12").exists())

    def test_teacher_can_import_ordering_question_excel(self):
        self.client.force_login(self.teacher)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(
            [
                "subject",
                "category",
                "question_type",
                "question_text",
                "difficulty_level",
                "points",
                "ordering_items",
            ]
        )
        worksheet.append(
            [
                "Matematika",
                "Aljabar",
                "ordering",
                "Urutkan langkah penyelesaian persamaan linear.",
                "medium",
                7,
                json.dumps(
                    [
                        {"text": "Pindahkan konstanta ke ruas kanan", "order": 1},
                        {"text": "Sederhanakan ruas sejenis", "order": 2},
                        {"text": "Bagi dengan koefisien variabel", "order": 3},
                    ]
                ),
            ]
        )
        payload = BytesIO()
        workbook.save(payload)
        upload = SimpleUploadedFile(
            "import_soal_ordering.xlsx",
            payload.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(reverse("question_import"), data={"import_file": upload})

        self.assertEqual(response.status_code, 200)
        created = Question.objects.get(question_text__icontains="persamaan linear")
        self.assertEqual(created.question_type, "ordering")
        self.assertEqual(
            list(created.ordering_items.order_by("correct_order").values_list("item_text", flat=True)),
            [
                "Pindahkan konstanta ke ruas kanan",
                "Sederhanakan ruas sejenis",
                "Bagi dengan koefisien variabel",
            ],
        )

    def test_teacher_can_import_matching_question_excel(self):
        self.client.force_login(self.teacher)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(
            [
                "subject",
                "category",
                "question_type",
                "question_text",
                "difficulty_level",
                "points",
                "matching_pairs",
            ]
        )
        worksheet.append(
            [
                "Matematika",
                "Aljabar",
                "matching",
                "Pasangkan bentuk dengan jumlah sisi.",
                "easy",
                6,
                json.dumps(
                    [
                        {"prompt": "Segitiga", "answer": "3", "order": 1},
                        {"prompt": "Persegi", "answer": "4", "order": 2},
                    ]
                ),
            ]
        )
        payload = BytesIO()
        workbook.save(payload)
        upload = SimpleUploadedFile(
            "import_soal_matching.xlsx",
            payload.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(reverse("question_import"), data={"import_file": upload})

        self.assertEqual(response.status_code, 200)
        created = Question.objects.get(question_text__icontains="bentuk dengan jumlah sisi")
        self.assertEqual(created.question_type, "matching")
        self.assertEqual(created.matching_pairs.count(), 2)
        self.assertTrue(created.matching_pairs.filter(prompt_text="Segitiga", answer_text="3").exists())

    def test_teacher_can_import_fill_in_blank_question_excel(self):
        self.client.force_login(self.teacher)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(
            [
                "subject",
                "category",
                "question_type",
                "question_text",
                "difficulty_level",
                "points",
                "blank_answers",
            ]
        )
        worksheet.append(
            [
                "Matematika",
                "Aljabar",
                "fill_in_blank",
                "Hasil 5 + 5 adalah {{1}}.",
                "easy",
                5,
                json.dumps(
                    {
                        "1": {
                            "accepted_answers": ["10", "sepuluh"],
                            "is_case_sensitive": False,
                            "blank_points": 5,
                        }
                    }
                ),
            ]
        )
        payload = BytesIO()
        workbook.save(payload)
        upload = SimpleUploadedFile(
            "import_soal_fill_blank.xlsx",
            payload.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(reverse("question_import"), data={"import_file": upload})

        self.assertEqual(response.status_code, 200)
        created = Question.objects.get(question_text__icontains="Hasil 5 + 5")
        self.assertEqual(created.question_type, "fill_in_blank")
        self.assertEqual(created.blank_answers.count(), 1)
        self.assertEqual(created.blank_answers.get(blank_number=1).accepted_answers, ["10", "sepuluh"])

    def test_question_import_template_includes_option_j_header(self):
        self.client.force_login(self.teacher)
        response = self.client.get(reverse("question_import_template"))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        worksheet = workbook.active
        headers = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))

        self.assertIn("option_j", headers)
        self.assertIn("audio_play_limit", headers)
        self.assertIn("video_play_limit", headers)
        self.assertIn("ordering_items", headers)
        self.assertIn("matching_pairs", headers)
        self.assertIn("blank_answers", headers)
        self.assertLess(headers.index("option_j"), headers.index("correct_option"))

    def test_question_import_history_returns_empty_when_log_table_unavailable(self):
        with patch(
            "apps.questions.services.QuestionImportLog.objects.select_related",
            side_effect=ProgrammingError("table missing"),
        ):
            history = get_question_import_history(self.teacher, limit=10)

        self.assertEqual(history, [])
