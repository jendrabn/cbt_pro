from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
import json
import re
import uuid

from django.conf import settings
from django.db import OperationalError, ProgrammingError, transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.accounts.models import UserActivityLog
from apps.subjects.models import Subject

from .models import (
    Question,
    QuestionAnswer,
    QuestionBlankAnswer,
    QuestionCategory,
    QuestionImportLog,
    QuestionMatchingPair,
    QuestionOption,
    QuestionOrderingItem,
    QuestionTag,
    QuestionTagRelation,
)
from .richtext import sanitize_optional_richtext_html, sanitize_richtext_html
from .services import (
    parse_tags,
    sync_question_answer,
    sync_question_blank_answers,
    sync_question_matching_pairs,
    sync_question_options,
    sync_question_ordering_items,
    sync_question_tags,
)


HEADER_ALIASES = {
    "mata_pelajaran": "subject",
    "tipe_soal": "question_type",
    "teks_soal": "question_text",
    "tingkat_kesulitan": "difficulty_level",
    "pembahasan": "explanation",
    "gambar_soal": "question_image_url",
    "batas_putar_audio": "audio_play_limit",
    "batas_putar_video": "video_play_limit",
    "batas_waktu_soal": "time_limit_seconds",
}
OPTION_LETTERS = [choice.value for choice in QuestionOption.OptionLetter]
BLANK_PLACEHOLDER_RE = re.compile(r"\{\{\s*(\d+)\s*\}\}")


@dataclass
class ImportResult:
    total_rows: int = 0
    success_count: int = 0
    skipped_count: int = 0
    errors: list[str] = field(default_factory=list)
    error_details: list[dict] = field(default_factory=list)
    skip_details: list[dict] = field(default_factory=list)
    import_log_id: str = ""


@dataclass
class PreparedQuestionRow:
    row_number: int
    question_id: uuid.UUID
    original_payload: dict
    subject: Subject
    category_name: str | None
    question_type: str
    question_text: str
    difficulty_level: str | None
    points: Decimal
    explanation: str | None
    checkbox_scoring: str
    question_image_url: str | None
    audio_play_limit: int | None
    video_play_limit: int | None
    allow_previous: bool
    allow_next: bool
    force_sequential: bool
    time_limit_seconds: int | None
    is_active: bool
    options: list[dict] = field(default_factory=list)
    ordering_items: list[dict] = field(default_factory=list)
    matching_pairs: list[dict] = field(default_factory=list)
    blank_answers: list[dict] = field(default_factory=list)
    answer_payload: dict | None = None
    tag_names: list[str] = field(default_factory=list)


def _normalize_header(value):
    if value is None:
        return ""
    normalized = re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")
    return HEADER_ALIASES.get(normalized, normalized)


def _normalize_question_type(value):
    raw = (value or "").strip().lower()
    mapping = {
        "multiple_choice": "multiple_choice",
        "pilihan_ganda": "multiple_choice",
        "pilihan ganda": "multiple_choice",
        "checkbox": "checkbox",
        "ordering": "ordering",
        "matching": "matching",
        "fill_in_blank": "fill_in_blank",
        "fill in blank": "fill_in_blank",
        "essay": "essay",
        "esai": "essay",
        "short_answer": "short_answer",
        "jawaban_singkat": "short_answer",
        "jawaban singkat": "short_answer",
    }
    if raw not in mapping:
        raise ValueError("Tipe soal tidak valid.")
    return mapping[raw]


def _normalize_difficulty(value):
    raw = (value or "").strip().lower()
    if not raw:
        return None
    mapping = {
        "easy": "easy",
        "mudah": "easy",
        "medium": "medium",
        "sedang": "medium",
        "hard": "hard",
        "sulit": "hard",
    }
    if raw not in mapping:
        raise ValueError("Tingkat kesulitan tidak valid.")
    return mapping[raw]


def _parse_bool(value, default=False):
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    raw = str(value).strip().lower()
    return raw in {"1", "true", "ya", "yes", "y"}


def _parse_decimal(value, default=Decimal("1.00")):
    if value in (None, ""):
        return default
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError):
        raise ValueError("Nilai poin tidak valid.")


def _parse_int(value, field_label):
    if value in (None, ""):
        return None
    try:
        output = int(value)
    except (TypeError, ValueError):
        raise ValueError(f"{field_label} harus berupa angka.")
    if output <= 0:
        raise ValueError(f"{field_label} harus lebih dari 0.")
    return output


def _build_subject_maps():
    active_subjects = Subject.objects.filter(is_active=True).only("id", "name", "code")
    subject_by_code = {subject.code.lower(): subject for subject in active_subjects if subject.code}
    subject_by_name = {subject.name.lower(): subject for subject in active_subjects if subject.name}
    return subject_by_code, subject_by_name


def _resolve_subject(subject_value, subject_by_code, subject_by_name):
    target = (subject_value or "").strip()
    if not target:
        raise ValueError("Mata pelajaran wajib diisi.")
    subject = subject_by_code.get(target.lower()) or subject_by_name.get(target.lower())
    if not subject:
        raise ValueError(f"Mata pelajaran '{target}' tidak ditemukan.")
    return subject


def _parse_ordering_items(value):
    if value in (None, ""):
        return []

    parsed_value = value
    if isinstance(value, str):
        raw_value = value.strip()
        if not raw_value:
            return []
        try:
            parsed_value = json.loads(raw_value)
        except json.JSONDecodeError as exc:
            raise ValueError("ordering_items harus berupa JSON array.") from exc

    if not isinstance(parsed_value, (list, tuple)):
        raise ValueError("ordering_items harus berupa JSON array.")

    normalized_items = []
    for index, item in enumerate(parsed_value, start=1):
        text_value = item
        sort_order = index
        if isinstance(item, dict):
            text_value = item.get("text") or item.get("item_text") or item.get("label") or ""
            sort_order = item.get("order") or item.get("correct_order") or index

        item_text = sanitize_richtext_html(text_value)
        if not item_text:
            continue

        try:
            normalized_order = int(sort_order)
        except (TypeError, ValueError):
            normalized_order = index
        normalized_items.append((normalized_order, item_text))

    normalized_items.sort(key=lambda item: item[0])
    return [item_text for _, item_text in normalized_items]


def _parse_matching_pairs(value):
    if value in (None, ""):
        return []

    parsed_value = value
    if isinstance(value, str):
        raw_value = value.strip()
        if not raw_value:
            return []
        try:
            parsed_value = json.loads(raw_value)
        except json.JSONDecodeError as exc:
            raise ValueError("matching_pairs harus berupa JSON array.") from exc

    if not isinstance(parsed_value, (list, tuple)):
        raise ValueError("matching_pairs harus berupa JSON array.")

    normalized_pairs = []
    for index, item in enumerate(parsed_value, start=1):
        if not isinstance(item, dict):
            raise ValueError("Setiap elemen matching_pairs harus berupa objek prompt/answer.")
        prompt_text = sanitize_richtext_html(item.get("prompt") or item.get("prompt_text"))
        answer_text = sanitize_richtext_html(item.get("answer") or item.get("answer_text"))
        if not prompt_text or not answer_text:
            continue
        pair_order = item.get("order") or item.get("pair_order") or index
        try:
            normalized_order = int(pair_order)
        except (TypeError, ValueError):
            normalized_order = index
        normalized_pairs.append(
            {
                "prompt_text": prompt_text,
                "answer_text": answer_text,
                "pair_order": normalized_order,
            }
        )

    normalized_pairs.sort(key=lambda item: item["pair_order"])
    return normalized_pairs


def _parse_blank_answers(value):
    if value in (None, ""):
        return []

    parsed_value = value
    if isinstance(value, str):
        raw_value = value.strip()
        if not raw_value:
            return []
        try:
            parsed_value = json.loads(raw_value)
        except json.JSONDecodeError as exc:
            raise ValueError("blank_answers harus berupa JSON object.") from exc

    if not isinstance(parsed_value, dict):
        raise ValueError("blank_answers harus berupa JSON object.")

    normalized_answers = []
    for raw_blank_number, payload in parsed_value.items():
        try:
            blank_number = int(raw_blank_number)
        except (TypeError, ValueError) as exc:
            raise ValueError("Kunci blank_answers harus berupa angka.") from exc

        accepted_answers = payload
        is_case_sensitive = False
        blank_points = None
        if isinstance(payload, dict):
            accepted_answers = payload.get("accepted_answers") or payload.get("answers") or []
            is_case_sensitive = bool(payload.get("is_case_sensitive"))
            blank_points = payload.get("blank_points")

        if not isinstance(accepted_answers, (list, tuple)):
            raise ValueError(f"blank_answers untuk {{{{{blank_number}}}}} harus berupa list jawaban.")

        normalized_list = [str(item).strip() for item in accepted_answers if str(item).strip()]
        if not normalized_list:
            continue

        normalized_answer = {
            "blank_number": blank_number,
            "accepted_answers": normalized_list,
            "is_case_sensitive": is_case_sensitive,
            "blank_points": None,
        }
        if blank_points not in (None, ""):
            normalized_answer["blank_points"] = _parse_decimal(blank_points, default=Decimal("0.00"))
        normalized_answers.append(normalized_answer)

    normalized_answers.sort(key=lambda item: item["blank_number"])
    return normalized_answers


def _extract_cleaned_payload(payload, question_type):
    option_values = {
        letter: (payload.get(f"option_{letter.lower()}") or "").strip()
        for letter in OPTION_LETTERS
    }
    active_options = {letter: text for letter, text in option_values.items() if text}
    correct_option = (payload.get("correct_option") or "").strip().upper()
    correct_options = []
    seen_correct_options = set()
    for item in str(payload.get("correct_options") or payload.get("correct_option") or "").split(","):
        normalized_item = item.strip().upper()
        if not normalized_item or normalized_item in seen_correct_options:
            continue
        seen_correct_options.add(normalized_item)
        correct_options.append(normalized_item)
    checkbox_scoring = (payload.get("checkbox_scoring") or "").strip().lower() or Question.CheckboxScoring.ALL_OR_NOTHING
    ordering_items = _parse_ordering_items(payload.get("ordering_items"))
    matching_pairs = _parse_matching_pairs(payload.get("matching_pairs"))
    blank_answers = _parse_blank_answers(payload.get("blank_answers"))
    blank_numbers = {item["blank_number"] for item in blank_answers}
    question_text_raw = str(payload.get("question_text") or "")

    if question_type == Question.QuestionType.MULTIPLE_CHOICE:
        if len(active_options) < 2:
            raise ValueError("Soal pilihan ganda wajib memiliki minimal 2 opsi.")
        if correct_option not in active_options:
            raise ValueError("Jawaban benar harus salah satu dari opsi yang terisi.")
        correct_options = []
        checkbox_scoring = Question.CheckboxScoring.ALL_OR_NOTHING
    elif question_type == Question.QuestionType.CHECKBOX:
        if len(active_options) < 2:
            raise ValueError("Soal checkbox wajib memiliki minimal 2 opsi.")
        valid_scoring = {choice[0] for choice in Question.CheckboxScoring.choices}
        if checkbox_scoring not in valid_scoring:
            raise ValueError("Metode penilaian checkbox tidak valid.")
        if len(correct_options) < 2:
            raise ValueError("Soal checkbox wajib memiliki minimal 2 jawaban benar.")
        invalid_correct = [letter for letter in correct_options if letter not in active_options]
        if invalid_correct:
            raise ValueError("Jawaban benar checkbox harus berasal dari opsi yang terisi.")
        correct_option = ""
        ordering_items = []
    elif question_type == Question.QuestionType.ORDERING:
        if len(ordering_items) < 2:
            raise ValueError("Soal ordering wajib memiliki minimal 2 item.")
        correct_option = ""
        correct_options = []
        checkbox_scoring = Question.CheckboxScoring.ALL_OR_NOTHING
        matching_pairs = []
        blank_answers = []
    elif question_type == Question.QuestionType.MATCHING:
        if len(matching_pairs) < 2:
            raise ValueError("Soal matching wajib memiliki minimal 2 pasangan.")
        correct_option = ""
        correct_options = []
        checkbox_scoring = Question.CheckboxScoring.ALL_OR_NOTHING
        ordering_items = []
        blank_answers = []
    elif question_type == Question.QuestionType.FILL_IN_BLANK:
        if not BLANK_PLACEHOLDER_RE.search(question_text_raw):
            raise ValueError("Soal fill in blank harus mengandung placeholder seperti {{1}}.")
        if not blank_answers:
            raise ValueError("Soal fill in blank wajib memiliki definisi blank_answers.")
        question_blank_numbers = {int(number) for number in BLANK_PLACEHOLDER_RE.findall(question_text_raw)}
        if question_blank_numbers != blank_numbers:
            raise ValueError("Definisi blank_answers harus cocok dengan placeholder {{n}} pada question_text.")
        correct_option = ""
        correct_options = []
        checkbox_scoring = Question.CheckboxScoring.ALL_OR_NOTHING
        ordering_items = []
        matching_pairs = []
    else:
        if not (payload.get("answer_text") or "").strip():
            raise ValueError("Kunci jawaban/rubrik wajib diisi untuk tipe soal non-PG.")
        correct_option = ""
        correct_options = []
        checkbox_scoring = Question.CheckboxScoring.ALL_OR_NOTHING
        ordering_items = []
        matching_pairs = []
        blank_answers = []

    cleaned_payload = {
        "question_type": question_type,
        "correct_option": correct_option,
        "correct_options": correct_options,
        "checkbox_scoring": checkbox_scoring,
        "ordering_items": ordering_items,
        "matching_pairs": matching_pairs,
        "blank_answers": blank_answers,
        "answer_text": (payload.get("answer_text") or "").strip(),
        "keywords": (payload.get("keywords") or "").strip(),
        "is_case_sensitive": _parse_bool(payload.get("is_case_sensitive"), default=False),
        "max_word_count": _parse_int(payload.get("max_word_count"), "Batas kata"),
        "tags": (payload.get("tags") or "").strip(),
    }
    for letter in OPTION_LETTERS:
        cleaned_payload[f"option_{letter.lower()}"] = option_values[letter]
    return cleaned_payload


@transaction.atomic
def _create_question_from_payload(payload, teacher):
    question_type = _normalize_question_type(payload.get("question_type"))
    question_text = sanitize_richtext_html(payload.get("question_text"))
    if not question_text:
        raise ValueError("Teks soal wajib diisi.")
    cleaned_data = _extract_cleaned_payload(payload, question_type)

    subject_by_code, subject_by_name = _build_subject_maps()
    subject = _resolve_subject(payload.get("subject"), subject_by_code, subject_by_name)
    category_name = (payload.get("category") or "").strip()
    category = None
    if category_name:
        category, _ = QuestionCategory.objects.get_or_create(name=category_name, defaults={"is_active": True})
    difficulty_level = _normalize_difficulty(payload.get("difficulty_level"))
    points = _parse_decimal(payload.get("points"))
    allow_previous = _parse_bool(payload.get("allow_previous"), default=True)
    allow_next = _parse_bool(payload.get("allow_next"), default=True)
    force_sequential = _parse_bool(payload.get("force_sequential"), default=False)
    if force_sequential:
        allow_previous = False

    question = Question.objects.create(
        created_by=teacher,
        subject=subject,
        category=category,
        question_type=question_type,
        question_text=question_text,
        question_image_url=(payload.get("question_image_url") or "").strip() or None,
        audio_play_limit=_parse_int(payload.get("audio_play_limit"), "Batas putar audio"),
        video_play_limit=_parse_int(payload.get("video_play_limit"), "Batas putar video"),
        points=points,
        difficulty_level=difficulty_level,
        explanation=sanitize_optional_richtext_html(payload.get("explanation")),
        checkbox_scoring=cleaned_data.get("checkbox_scoring") or Question.CheckboxScoring.ALL_OR_NOTHING,
        allow_previous=allow_previous,
        allow_next=allow_next,
        force_sequential=force_sequential,
        time_limit_seconds=_parse_int(payload.get("time_limit_seconds"), "Batas waktu soal"),
        is_active=_parse_bool(payload.get("is_active"), default=True),
    )

    sync_question_options(question, cleaned_data)
    sync_question_ordering_items(question, cleaned_data)
    sync_question_matching_pairs(question, cleaned_data)
    sync_question_blank_answers(question, cleaned_data)
    sync_question_answer(question, cleaned_data)
    sync_question_tags(question, cleaned_data)
    return question


def _append_error(result: ImportResult, row_number: int, error: str):
    result.errors.append(f"Baris Excel {row_number}: {error}")
    result.error_details.append({"row": row_number, "error": error})


def _append_skip(result: ImportResult, row_number: int, reason: str):
    result.skipped_count += 1
    result.skip_details.append({"row": row_number, "reason": reason})


def _chunked(items, chunk_size: int):
    chunk_size = max(1, chunk_size)
    for start in range(0, len(items), chunk_size):
        yield items[start:start + chunk_size]


def _ensure_categories(prepared_rows: list[PreparedQuestionRow], category_cache: dict[str, QuestionCategory]):
    missing_names = []
    for row in prepared_rows:
        if row.category_name and row.category_name not in category_cache:
            missing_names.append(row.category_name)

    if not missing_names:
        return

    unique_names = list(dict.fromkeys(missing_names))
    existing_categories = QuestionCategory.objects.filter(name__in=unique_names)
    for category in existing_categories:
        category_cache[category.name] = category

    new_categories = [
        QuestionCategory(name=name, is_active=True)
        for name in unique_names
        if name not in category_cache
    ]
    if new_categories:
        QuestionCategory.objects.bulk_create(
            new_categories,
            batch_size=getattr(settings, "QUESTION_IMPORT_CHUNK_SIZE", 200),
        )
        for category in new_categories:
            category_cache[category.name] = category


def _ensure_tags(prepared_rows: list[PreparedQuestionRow], tag_cache: dict[str, QuestionTag]):
    missing_names = []
    for row in prepared_rows:
        for tag_name in row.tag_names:
            if tag_name not in tag_cache:
                missing_names.append(tag_name)

    if not missing_names:
        return

    unique_names = list(dict.fromkeys(missing_names))
    existing_tags = QuestionTag.objects.filter(name__in=unique_names)
    for tag in existing_tags:
        tag_cache[tag.name] = tag

    new_tags = [QuestionTag(name=name) for name in unique_names if name not in tag_cache]
    if new_tags:
        QuestionTag.objects.bulk_create(
            new_tags,
            batch_size=getattr(settings, "QUESTION_IMPORT_CHUNK_SIZE", 200),
            ignore_conflicts=True,
        )
        for tag in QuestionTag.objects.filter(name__in=unique_names):
            tag_cache[tag.name] = tag


def _prepare_question_row(row_index, row_values, headers, subject_by_code, subject_by_name):
    row = {headers[idx]: row_values[idx] if idx < len(row_values) else None for idx in range(len(headers))}
    normalized = {key: ("" if value is None else str(value)) for key, value in row.items()}

    for numeric_key in ["points", "time_limit_seconds", "max_word_count"]:
        if numeric_key in row:
            normalized[numeric_key] = row[numeric_key]

    question_type = _normalize_question_type(normalized.get("question_type"))
    question_text = sanitize_richtext_html(normalized.get("question_text"))
    if not question_text:
        raise ValueError("Teks soal wajib diisi.")

    subject = _resolve_subject(normalized.get("subject"), subject_by_code, subject_by_name)
    difficulty_level = _normalize_difficulty(normalized.get("difficulty_level"))
    points = _parse_decimal(normalized.get("points"))
    allow_previous = _parse_bool(normalized.get("allow_previous"), default=True)
    allow_next = _parse_bool(normalized.get("allow_next"), default=True)
    force_sequential = _parse_bool(normalized.get("force_sequential"), default=False)
    if force_sequential:
        allow_previous = False

    cleaned_data = _extract_cleaned_payload(normalized, question_type)
    options = []
    ordering_items = []
    matching_pairs = []
    blank_answers = []
    if question_type in {Question.QuestionType.MULTIPLE_CHOICE, Question.QuestionType.CHECKBOX}:
        display_order = 1
        for letter in OPTION_LETTERS:
            text = sanitize_richtext_html(cleaned_data.get(f"option_{letter.lower()}"))
            if not text:
                continue
            options.append(
                {
                    "option_letter": letter,
                    "option_text": text,
                    "is_correct": (
                        cleaned_data.get("correct_option") == letter
                        if question_type == Question.QuestionType.MULTIPLE_CHOICE
                        else letter in set(cleaned_data.get("correct_options") or [])
                    ),
                    "display_order": display_order,
                }
            )
            display_order += 1
    elif question_type == Question.QuestionType.ORDERING:
        ordering_items = [
            {
                "item_text": item_text,
                "correct_order": index,
            }
            for index, item_text in enumerate(cleaned_data.get("ordering_items") or [], start=1)
        ]
    elif question_type == Question.QuestionType.MATCHING:
        matching_pairs = list(cleaned_data.get("matching_pairs") or [])
    elif question_type == Question.QuestionType.FILL_IN_BLANK:
        blank_answers = list(cleaned_data.get("blank_answers") or [])

    answer_payload = None
    if question_type not in {
        Question.QuestionType.MULTIPLE_CHOICE,
        Question.QuestionType.CHECKBOX,
        Question.QuestionType.ORDERING,
        Question.QuestionType.MATCHING,
        Question.QuestionType.FILL_IN_BLANK,
    }:
        answer_payload = {
            "answer_text": cleaned_data.get("answer_text", ""),
            "keywords": parse_tags(cleaned_data.get("keywords")),
            "is_case_sensitive": bool(cleaned_data.get("is_case_sensitive")) if question_type == "short_answer" else False,
            "max_word_count": cleaned_data.get("max_word_count"),
        }

    return PreparedQuestionRow(
        row_number=row_index,
        question_id=uuid.uuid4(),
        original_payload=normalized,
        subject=subject,
        category_name=(normalized.get("category") or "").strip() or None,
        question_type=question_type,
        question_text=question_text,
        difficulty_level=difficulty_level,
        points=points,
        explanation=sanitize_optional_richtext_html(normalized.get("explanation")),
        checkbox_scoring=cleaned_data.get("checkbox_scoring") or Question.CheckboxScoring.ALL_OR_NOTHING,
        question_image_url=(normalized.get("question_image_url") or "").strip() or None,
        audio_play_limit=_parse_int(normalized.get("audio_play_limit"), "Batas putar audio"),
        video_play_limit=_parse_int(normalized.get("video_play_limit"), "Batas putar video"),
        allow_previous=allow_previous,
        allow_next=allow_next,
        force_sequential=force_sequential,
        time_limit_seconds=_parse_int(normalized.get("time_limit_seconds"), "Batas waktu soal"),
        is_active=_parse_bool(normalized.get("is_active"), default=True),
        options=options,
        ordering_items=ordering_items,
        matching_pairs=matching_pairs,
        blank_answers=blank_answers,
        answer_payload=answer_payload,
        tag_names=parse_tags(cleaned_data.get("tags")),
    )


def _process_question_chunk(
    prepared_rows: list[PreparedQuestionRow],
    teacher,
    result: ImportResult,
    category_cache: dict[str, QuestionCategory],
    tag_cache: dict[str, QuestionTag],
):
    if not prepared_rows:
        return

    try:
        with transaction.atomic():
            _ensure_categories(prepared_rows, category_cache)
            _ensure_tags(prepared_rows, tag_cache)

            question_objects = []
            option_objects = []
            ordering_item_objects = []
            matching_pair_objects = []
            blank_answer_objects = []
            answer_objects = []
            relation_objects = []

            for row in prepared_rows:
                category = category_cache.get(row.category_name) if row.category_name else None
                question_objects.append(
                    Question(
                        id=row.question_id,
                        created_by=teacher,
                        subject_id=row.subject.id,
                        category_id=category.id if category else None,
                        question_type=row.question_type,
                        question_text=row.question_text,
                        question_image_url=row.question_image_url,
                        audio_play_limit=row.audio_play_limit,
                        video_play_limit=row.video_play_limit,
                        points=row.points,
                        difficulty_level=row.difficulty_level,
                        explanation=row.explanation,
                        checkbox_scoring=row.checkbox_scoring,
                        allow_previous=row.allow_previous,
                        allow_next=row.allow_next,
                        force_sequential=row.force_sequential,
                        time_limit_seconds=row.time_limit_seconds,
                        is_active=row.is_active,
                    )
                )
                for option in row.options:
                    option_objects.append(
                        QuestionOption(
                            question_id=row.question_id,
                            option_letter=option["option_letter"],
                            option_text=option["option_text"],
                            is_correct=option["is_correct"],
                            display_order=option["display_order"],
                        )
                    )
                for ordering_item in row.ordering_items:
                    ordering_item_objects.append(
                        QuestionOrderingItem(
                            question_id=row.question_id,
                            item_text=ordering_item["item_text"],
                            correct_order=ordering_item["correct_order"],
                        )
                    )
                for matching_pair in row.matching_pairs:
                    matching_pair_objects.append(
                        QuestionMatchingPair(
                            question_id=row.question_id,
                            prompt_text=matching_pair["prompt_text"],
                            answer_text=matching_pair["answer_text"],
                            pair_order=matching_pair["pair_order"],
                        )
                    )
                for blank_answer in row.blank_answers:
                    blank_answer_objects.append(
                        QuestionBlankAnswer(
                            question_id=row.question_id,
                            blank_number=blank_answer["blank_number"],
                            accepted_answers=blank_answer["accepted_answers"],
                            is_case_sensitive=blank_answer.get("is_case_sensitive", False),
                            blank_points=blank_answer.get("blank_points"),
                        )
                    )
                if row.answer_payload:
                    answer_objects.append(
                        QuestionAnswer(
                            question_id=row.question_id,
                            answer_text=row.answer_payload["answer_text"],
                            keywords=row.answer_payload["keywords"],
                            is_case_sensitive=row.answer_payload["is_case_sensitive"],
                            max_word_count=row.answer_payload["max_word_count"],
                        )
                    )
                for tag_name in row.tag_names:
                    relation_objects.append(
                        QuestionTagRelation(
                            question_id=row.question_id,
                            tag_id=tag_cache[tag_name].id,
                        )
                    )

            Question.objects.bulk_create(
                question_objects,
                batch_size=getattr(settings, "QUESTION_IMPORT_CHUNK_SIZE", 200),
            )
            if option_objects:
                QuestionOption.objects.bulk_create(
                    option_objects,
                    batch_size=getattr(settings, "QUESTION_IMPORT_CHUNK_SIZE", 200),
                )
            if ordering_item_objects:
                QuestionOrderingItem.objects.bulk_create(
                    ordering_item_objects,
                    batch_size=getattr(settings, "QUESTION_IMPORT_CHUNK_SIZE", 200),
                )
            if matching_pair_objects:
                QuestionMatchingPair.objects.bulk_create(
                    matching_pair_objects,
                    batch_size=getattr(settings, "QUESTION_IMPORT_CHUNK_SIZE", 200),
                )
            if blank_answer_objects:
                QuestionBlankAnswer.objects.bulk_create(
                    blank_answer_objects,
                    batch_size=getattr(settings, "QUESTION_IMPORT_CHUNK_SIZE", 200),
                )
            if answer_objects:
                QuestionAnswer.objects.bulk_create(
                    answer_objects,
                    batch_size=getattr(settings, "QUESTION_IMPORT_CHUNK_SIZE", 200),
                )
            if relation_objects:
                QuestionTagRelation.objects.bulk_create(
                    relation_objects,
                    batch_size=getattr(settings, "QUESTION_IMPORT_CHUNK_SIZE", 200),
                    ignore_conflicts=True,
                )

        result.success_count += len(prepared_rows)
    except Exception:
        for row in prepared_rows:
            try:
                with transaction.atomic():
                    _create_question_from_payload(row.original_payload, teacher)
                result.success_count += 1
            except Exception as exc:
                _append_error(result, row.row_number, str(exc))


def _create_import_log(uploaded_file, teacher):
    try:
        return QuestionImportLog.objects.create(
            imported_by=teacher,
            original_filename=getattr(uploaded_file, "name", "unknown.xlsx"),
            file_size_kb=getattr(uploaded_file, "size", 0) // 1024,
            status="processing",
            started_at=timezone.now(),
        )
    except (ProgrammingError, OperationalError):
        return None


def _save_import_log(import_log, **fields):
    if import_log is None:
        return
    for key, value in fields.items():
        setattr(import_log, key, value)
    try:
        import_log.save()
    except (ProgrammingError, OperationalError):
        return


def import_questions_from_excel(uploaded_file, teacher):
    result = ImportResult()
    import_log = _create_import_log(uploaded_file, teacher)
    if import_log is not None:
        result.import_log_id = str(import_log.id)

    try:
        try:
            workbook = load_workbook(uploaded_file, data_only=True, read_only=True)
        except Exception as exc:
            result.errors.append(f"Gagal membaca file Excel: {exc}")
            _save_import_log(
                import_log,
                status="failed",
                total_failed=1,
                error_details=[{"row": 0, "error": str(exc)}],
                finished_at=timezone.now(),
            )
            return result

        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            result.errors.append("File Excel kosong.")
            _save_import_log(
                import_log,
                status="failed",
                total_failed=1,
                error_details=[{"row": 0, "error": "File Excel kosong."}],
                finished_at=timezone.now(),
            )
            return result

        headers = [_normalize_header(value) for value in header_row]
        required_headers = {"subject", "question_type", "question_text"}
        if not required_headers.issubset(set(headers)):
            error = "Header wajib minimal: subject, question_type, question_text."
            result.errors.append(error)
            _save_import_log(
                import_log,
                status="failed",
                total_failed=1,
                error_details=[{"row": 0, "error": error}],
                finished_at=timezone.now(),
            )
            return result

        subject_by_code, subject_by_name = _build_subject_maps()
        category_cache: dict[str, QuestionCategory] = {}
        tag_cache: dict[str, QuestionTag] = {}
        pending_rows: list[PreparedQuestionRow] = []

        max_rows = getattr(settings, "QUESTION_IMPORT_MAX_ROWS", 10000)
        chunk_size = getattr(settings, "QUESTION_IMPORT_CHUNK_SIZE", 200)

        for row_index, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in row_values):
                continue

            result.total_rows += 1
            if result.total_rows > max_rows:
                _append_error(result, row_index, f"Maksimal {max_rows} baris data.")
                break

            try:
                prepared_row = _prepare_question_row(
                    row_index,
                    row_values,
                    headers,
                    subject_by_code,
                    subject_by_name,
                )
                pending_rows.append(prepared_row)
            except Exception as exc:
                _append_error(result, row_index, str(exc))
                continue

            if len(pending_rows) >= chunk_size:
                _process_question_chunk(pending_rows, teacher, result, category_cache, tag_cache)
                pending_rows = []

        if pending_rows:
            _process_question_chunk(pending_rows, teacher, result, category_cache, tag_cache)

        _save_import_log(
            import_log,
            total_rows=result.total_rows,
            total_created=result.success_count,
            total_skipped=result.skipped_count,
            total_failed=len(result.error_details),
            error_details=result.error_details,
            skip_details=result.skip_details,
            status="completed",
            finished_at=timezone.now(),
        )

        UserActivityLog.objects.create(
            user=teacher,
            action="import_soal",
            description=(
                f"Import soal dari file {getattr(import_log, 'original_filename', getattr(uploaded_file, 'name', 'unknown.xlsx'))}: "
                f"{result.success_count} berhasil, {len(result.error_details)} gagal."
            ),
        )
        return result
    except Exception as exc:
        result.errors.append(f"Terjadi kesalahan saat import soal: {exc}")
        _save_import_log(
            import_log,
            total_rows=result.total_rows,
            total_created=result.success_count,
            total_skipped=result.skipped_count,
            total_failed=len(result.error_details) + 1,
            error_details=result.error_details + [{"row": 0, "error": str(exc)}],
            skip_details=result.skip_details,
            status="failed",
            finished_at=timezone.now(),
        )
        return result


def export_template_headers():
    return [
        "subject",
        "category",
        "question_type",
        "question_text",
        "difficulty_level",
        "points",
        "explanation",
        "question_image_url",
        "audio_play_limit",
        "video_play_limit",
        "allow_previous",
        "allow_next",
        "force_sequential",
        "time_limit_seconds",
        *[f"option_{letter.lower()}" for letter in OPTION_LETTERS],
        "correct_option",
        "correct_options",
        "checkbox_scoring",
        "ordering_items",
        "matching_pairs",
        "blank_answers",
        "answer_text",
        "keywords",
        "is_case_sensitive",
        "max_word_count",
        "tags",
        "is_active",
    ]
